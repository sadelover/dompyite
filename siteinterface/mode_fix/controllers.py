# -*- coding: utf-8 -*-
from siteinterface.mode_fix import bp_fix
from flask import request, jsonify
from siteinterface.BEOPSqliteAccess import *
from siteinterface.BEOPDataAccess import BEOPDataAccess
import logging
from PyQt4.QtGui import QTextDocument, QPrinter, QApplication
import sys
import os
from bs4 import BeautifulSoup
from docx import Document
from openpyxl import Workbook
from docx.enum.style import WD_STYLE_TYPE
from docx.shared import Pt
from docx.oxml.ns import qn
from .utils import *
from zipfile import is_zipfile


@bp_fix.route("/add", methods=["POST"])
def add():
    try:
        rcv = request.get_json()
        pageId = rcv.get("pageId") if rcv.get("pageId") is not None else None
        importance = rcv.get("importance") if rcv.get("importance") is not None else None
        urgent = rcv.get("urgent") if rcv.get("urgent") is not None else None
        content = rcv.get("content") if rcv.get("content") is not None else None
        reportUser = rcv.get("reportUser") if rcv.get("reportUser") is not None else None
        energyEffects = rcv.get("energyEffects") if rcv.get("energyEffects") is not None else None
        solveUser = rcv.get("solveUser") if rcv.get("solveUser") is not None else ""
        x = rcv.get("x") if rcv.get("x") is not None else None
        y = rcv.get("y") if rcv.get("y") is not None else None
        visible = rcv.get("visible") if rcv.get("visible") is not None else None
        reportTitle = rcv.get("title") if rcv.get("title") is not None else None

        if pageId is None:
            return jsonify(dict(err=1, msg="pageId不能为空"))
        if x is None:
            return jsonify(dict(err=1, msg="x不能为空"))
        if y is None:
            return jsonify(dict(err=1, msg="y不能为空"))
        if importance is None:
            return jsonify(dict(err=1, msg="重要程度不能为空"))
        if urgent is None:
            return jsonify(dict(err=1, msg="紧急程度不能为空"))
        if content is None:
            return jsonify(dict(err=1, msg="报修内容不能为空"))
        if reportUser is None:
            return jsonify(dict(err=1, msg="reportUser不能为空"))
        if reportTitle is None:
            return jsonify(dict(err=1, msg="标题不能为空"))

        if not isinstance(pageId, str):
            return jsonify(dict(err=1, msg="pageId必须为字符串"))
        if not isinstance(x, int):
            return jsonify(dict(err=1, msg="x必须为整数"))
        if not isinstance(y, int):
            return jsonify(dict(err=1, msg="y必须为整数"))
        if not isinstance(importance, int):
            return jsonify(dict(err=1, msg="importance必须为整数"))
        if not isinstance(urgent, int):
            return jsonify(dict(err=1, msg="urgent必须为整数"))
        if not isinstance(content, str):
            return jsonify(dict(err=1, msg="content必须为字符串"))
        if not isinstance(reportUser, str):
            return jsonify(dict(err=1, msg="reportUser必须为字符串"))
        if not isinstance(solveUser, str):
            return jsonify(dict(err=1, msg="solveUser必须为字符串"))
        if not isinstance(reportTitle, str):
            return jsonify(dict(err=1, msg="reportTitle必须为字符串"))
        if len(reportTitle) > 30:
            return jsonify(dict(err=1, msg="标题字数不能超过30"))

        if energyEffects is not None:
            if not isinstance(energyEffects, int):
                return jsonify(dict(err=1, msg="energyEffects必须为整数"))

        if visible is not None:
            if not isinstance(visible, int):
                return jsonify(dict(err=1, msg="visible必须为整数"))

        maxId = BEOPDataAccess.getInstance().getMaxIdInTable("fix")
        fixId = maxId + 1
        reportTime = datetime.now()

        bSuc = BEOPDataAccess.getInstance().createFix(fixId, reportTime, importance, urgent, content, reportUser, solveUser, energyEffects, x, y, visible, pageId, reportTitle)

        if not bSuc:
            return jsonify(dict(err=1, msg="创建失败", data=False))
        return jsonify(dict(err=0, msg="创建成功", data=True))

    except Exception as e:
        strLog = "ERROR in /fix/add: %s" % e.__str__()
        logging.error(strLog)

        return jsonify(dict(err=1, msg="创建失败：%s" % e.__str__(), data=False))

@bp_fix.route("/remove", methods=["POST"])
def remove():
    try:
        rcv = request.get_json()
        fixId = rcv.get("fixId") if rcv.get("fixId") is not None else None

        if fixId is None:
            return jsonify(dict(err=1, msg="fixId不能为空"))
        if not isinstance(fixId, int):
            return jsonify(dict(err=1, msg="fixId必须为整数"))

        bSuc = BEOPDataAccess.getInstance().removeFix(fixId)

        if not bSuc:
            return jsonify(dict(err=1, msg="删除失败", data=False))
        return jsonify(dict(err=0, msg="删除成功", data=True))

    except Exception as e:
        strLog = "ERROR in /fix/remove: %s" % e.__str__()
        logging.error(strLog)

        return jsonify(dict(err=1, msg="删除失败：%s" % e.__str__(), data=False))

@bp_fix.route("/modify", methods=["POST"])
def modify():
    try:
        rcv = request.get_json()
        fixId = rcv.get("fixId") if rcv.get("fixId") is not None else None
        reportTime = rcv.get("reportTime") if rcv.get("reportTime") is not None else None
        importance = rcv.get("importance") if rcv.get("importance") is not None else None
        urgent = rcv.get("urgent") if rcv.get("urgent") is not None else None
        content = rcv.get("content") if rcv.get("content") is not None else None
        result = rcv.get("result") if rcv.get("result") is not None else None
        closeTime = rcv.get("closeTime") if rcv.get("closeTime") is not None else None
        reportUser = rcv.get("reportUser") if rcv.get("reportUser") is not None else None
        solveUser = rcv.get("solveUser") if rcv.get("solveUser") is not None else None
        energyEffects = rcv.get("energyEffects") if rcv.get("energyEffects") is not None else None
        x = rcv.get("x") if rcv.get("x") is not None else None
        y = rcv.get("y") if rcv.get("y") is not None else None
        visible = rcv.get("visible") if rcv.get("visible") is not None else None
        reportTitle = rcv.get("title") if rcv.get("title") is not None else None

        if fixId is None:
            return jsonify(dict(err=1, msg="fixId不能为空"))
        if not isinstance(fixId, int):
            return jsonify(dict(err=1, msg="fixId必须为整数"))

        if reportTime is not None:
            if not isValidDate(reportTime, "%Y-%m-%d %H:%M:%S"):
                return jsonify(dict(err=1, msg="reportTime格式有误"))

        if importance is not None:
            if not isinstance(importance, int):
                return jsonify(dict(err=1, msg="importance必须为整数"))

        if urgent is not None:
            if not isinstance(urgent, int):
                return jsonify(dict(err=1, msg="urgent必须为整数"))

        if content is not None:
            if not isinstance(content, str):
                return jsonify(dict(err=1, msg="报修内容必须为字符串"))

        if result is not None:
            if not isinstance(result, int):
                return jsonify(dict(err=1, msg="result必须为整数"))

        if closeTime is not None:
            if not isValidDate(closeTime, "%Y-%m-%d %H:%M:%S"):
                return jsonify(dict(err=1, msg="closeTime格式有误"))

        if reportUser is not None:
            if not isinstance(reportUser, str):
                return jsonify(dict(err=1, msg="reportUser必须为字符串"))

        if solveUser is not None:
            if not isinstance(solveUser, str):
                return jsonify(dict(err=1, msg="solveUser必须为字符串"))

        if energyEffects is not None:
            if not isinstance(energyEffects, int):
                return jsonify(dict(err=1, msg="energyEffects必须为整数"))

        if x is not None:
            if not isinstance(x, int):
                return jsonify(dict(err=1, msg="x必须为整数"))

        if y is not None:
            if not isinstance(y, int):
                return jsonify(dict(err=1, msg="y必须为整数"))

        if visible is not None:
            if not isinstance(visible, int):
                return jsonify(dict(err=1, msg="visible必须为整数"))

        if reportTitle is None:
            return jsonify(dict(err=1, msg="标题不能为空"))
        if not isinstance(reportTitle, str):
            return jsonify(dict(err=1, msg="reportTitle必须为字符串"))
        if len(reportTitle) > 30:
            return jsonify(dict(err=1, msg="标题字数不能超过30"))

        if not fixIdExists(fixId):
            return jsonify(dict(err=1, msg="fixId不存在"))

        paramDict = dict(reportTime=reportTime,
                         importance=importance,
                         urgent=urgent,
                         content=content,
                         result=result,
                         closeTime=closeTime,
                         reportUser=reportUser,
                         solveUser=solveUser,
                         energyEffects=energyEffects,
                         x=x,
                         y=y,
                         visible=visible,
                         title=reportTitle)

        paramDictToReplace = {}
        for key in paramDict.keys():
            if paramDict.get(key) is not None:
                paramDictToReplace[key] = paramDict.get(key)

        if not paramDictToReplace:
            return jsonify(dict(err=0, msg="未发现任何修改项", data=True))

        bSuc = BEOPDataAccess.getInstance().modifyFix(fixId, paramDictToReplace)
        if not bSuc:
            return jsonify(dict(err=1, msg="修改失败", data=False))

        return jsonify(dict(err=0, msg="修改成功", data=True))

    except Exception as e:
        strLog = "ERROR in /fix/modify: %s" % e.__str__()
        logging.error(strLog)
        return jsonify(dict(err=1, msg=strLog, data=False))

@bp_fix.route("/getById", methods=["POST"])
def get_by_id():
    try:
        rcv = request.get_json()
        fixId = rcv.get("fixId") if rcv.get("fixId") is not None else None

        if fixId is None:
            return jsonify(dict(err=1, msg="fixId不能为空"))
        if not isinstance(fixId, int):
            return jsonify(dict(err=1, msg="fixId必须为整数"))

        if not fixIdExists(fixId):
            return jsonify(dict(err=1, msg="fixId不存在"))

        arr = BEOPDataAccess.getInstance().getFixById(fixId)

        if not arr:
            return jsonify(dict(err=1, msg="获取失败", data={}))
        return jsonify(dict(err=0, msg="获取成功", data=arr))

    except Exception as e:
        strLog = "ERROR in /fix/getById: %s" % e.__str__()
        logging.error(strLog)
        return jsonify(dict(err=1, msg=strLog, data={}))


@bp_fix.route("/insertImage", methods=["POST"])
def insert_image():
    try:
        fileList = []
        for idx in range(1, 10):
            file = request.files.get("file%02d" % idx)
            if file is None or not isinstance(file.content_type, str):
                continue
            fileList.append(file)

        if not len(fileList):
            return jsonify(dict(err=1, msg="未发现上传的图片"))

        for file in fileList:

            strFileName = file.filename

            extension = os.path.splitext(strFileName)[1]
            if not extension in [".jpg", ".png", ".jpeg", ".bmp", ".JPG", ".PNG", ".JPEG", ".BMP"]:
                return jsonify(dict(err=1, msg="只能上传图片文件（支持jpg, png, jpeg, bmp格式）"))

        dResult = saveImageToLocal(fileList)

        return jsonify(dResult)

    except Exception as e:
        strLog = "ERROR in /fix/insertImage: %s" % e.__str__()
        logging.error(strLog)
        return jsonify(dict(err=1, msg="上传失败", data=""))


@bp_fix.route("/downloadFixContentInDocx", methods=["POST"])
def download_fix_content_in_docx():
    try:
        rcv = request.get_json()
        strTimeFormat = "%Y-%m-%d %H:%M:%S"
        levelDict = {0: "低", 1: "中", 2: "高"}
        resultDict = {0: "未解决", 1: "已解决"}

        fixId = rcv.get("fixId") if rcv.get("fixId") is not None else None
        if fixId is None:
            return jsonify(dict(err=1, msg="fixId不能为空", data=""))

        if not fixIdExists(fixId):
            return jsonify(dict(err=1, msg="该报修不存在", data=""))

        contentTuple = BEOPDataAccess.getInstance().getFixContentById(fixId)

        if not contentTuple:
            return jsonify(dict(err=1, msg="该报修无内容或获取内容失败", data=""))
        if not len(contentTuple):
            return jsonify(dict(err=1, msg="该报修无内容", data=""))

        reportTime = contentTuple[0].strftime(strTimeFormat)
        importance = levelDict.get(contentTuple[1], "低")
        urgent = levelDict.get(contentTuple[2], "低")
        strContent = contentTuple[3]
        result = resultDict.get(contentTuple[4], "无")
        closeTime = contentTuple[5].strftime(strTimeFormat) if contentTuple[5] > contentTuple[0] else "无"
        reportUser = contentTuple[6]
        solveUser = contentTuple[7] if len(contentTuple[7]) > 0 else "无"
        energyEffects = levelDict.get(contentTuple[8], "低")
        reportTitle = contentTuple[9]

        parList = process_fix_content(strContent)

        templateDocx = os.path.join(app.static_folder, "files", "fixReport.docx")
        if not os.path.exists(templateDocx):
            return jsonify(dict(err=1, msg="模板文件不存在，下载失败", data=""))
        if not is_zipfile(templateDocx):
            return jsonify(dict(err=1, msg="模板文件格式有误，下载失败", data=""))

        # 准备docx文件路径
        strNow = datetime.now().strftime("%Y%m%d%H%M%S")
        filesDir = os.path.join(app.static_folder, "files")
        if not os.path.exists(filesDir):
            os.mkdir(filesDir)

        destFileName = "fixReportDocx_{0}_{1}.docx".format(fixId, strNow)
        destFilePath = os.path.join(filesDir, destFileName)
        if os.path.exists(destFilePath):
            os.remove(destFilePath)

        bSucceed = genReportDocx(templateDocx, destFilePath, reportTitle, reportTime, reportUser, urgent, importance, energyEffects, solveUser, result, closeTime, parList)

        if not bSucceed:
            return jsonify(dict(err=1, msg="下载失败", data=""))

        return jsonify(dict(err=0, msg="下载成功", data=destFileName))

    except Exception as e:
        strLog = "ERROR in /fix/downloadFixContentInDocx: %s" % e.__str__()
        logging.error(strLog)
        return jsonify(dict(err=1, msg="下载失败", data=""))


@bp_fix.route("/downloadFixContentInExcel", methods=["POST"])
def download_fix_content_in_excel():
    try:
        rcv = request.get_json()
        nResult = rcv.get("result") if rcv.get("result") is not None else None
        if nResult is not None:
            if not isinstance(nResult, int):
                return jsonify(dict(err=1, msg="结果类型必须为整数", data=""))
            if not nResult in [0, 1, 2]:
                return jsonify(dict(err=1, msg="结果类型只支持0, 1, 2", data=""))

        dataList = BEOPDataAccess.getInstance().getFixContentByResult(nResult)

        if not dataList:
            return jsonify(dict(err=1, msg="获取失败或无内容", data=""))

        levelDict = {0: "低", 1: "中", 2: "高"}
        tableHeader = ["序号", "报告时间", "标题", "内容", "紧急程度", "重要程度", "节能影响程度", "报告人"]
        infoList = []
        for idx, data in enumerate(dataList):
            strTime = data[0].strftime("%Y-%m-%d %H:%M:%S")
            strContent = htmlToStr(data[1])
            urgent = levelDict.get(data[2], "低")
            importance = levelDict.get(data[3], "低")
            energyEffects = levelDict.get(data[4], "低")
            reportUser = data[5]
            reportTitle = data[6]
            tempList = [idx+1, strTime, reportTitle, strContent, urgent, importance, energyEffects, reportUser]
            infoList.append(tempList)

        if not len(infoList):
            return jsonify(dict(err=1, msg="无内容", data=""))

        strNow = datetime.now().strftime("%Y%m%d%H%M%S")

        filesDir = os.path.join(app.static_folder, "files")
        if not os.path.exists(filesDir):
            os.mkdir(filesDir)

        destExcelName = "fixReportExcel_{0}.xlsx".format(strNow)
        destExcelPath = os.path.join(filesDir, destExcelName)

        if os.path.exists(destExcelPath):
            os.remove(destExcelPath)

        book = Workbook()
        sheet = book.create_sheet("报修内容", 0)

        # 写表头
        for idx, item in enumerate(tableHeader):
            sheet.cell(row=1, column=idx+1, value=item)

        # 写表内容
        for nInfo, info in enumerate(infoList):
            for idx, item in enumerate(info):
                sheet.cell(row=nInfo+2, column=idx+1, value=item)

        book.save(destExcelPath)

        return jsonify(dict(err=0, msg="下载成功", data=destExcelName))

    except Exception as e:
        strLog = "ERROR in /fix/downloadFixContentInExcel: %s" % e.__str__()
        logging.error(strLog)
        return jsonify(dict(err=1, msg="下载失败", data=""))


@bp_fix.route("/getByPeriod", methods=["POST"])
def get_by_period():
    try:
        rcv = request.get_json()
        strTimeFormat = "%Y-%m-%d %H:%M:%S"
        strTimeFrom = rcv.get("timeFrom") if rcv.get("timeFrom") else None
        strTimeTo = rcv.get("timeTo") if rcv.get("timeTo") else None

        if strTimeFrom is None or strTimeTo is None:
            return jsonify(dict(err=1, msg="开始时间和结束时间不能为空", data=[]))

        if not isValidDate(strTimeFrom, strTimeFormat) or not isValidDate(strTimeTo, strTimeFormat):
            return jsonify(dict(err=1, msg="时间格式有误", data=[]))

        if datetime.strptime(strTimeFrom, strTimeFormat) > datetime.strptime(strTimeTo, strTimeFormat):
            return jsonify(dict(err=1, msg="开始时间不能大于结束时间", data=[]))
        if datetime.strptime(strTimeFrom, strTimeFormat) == datetime.strptime(strTimeTo, strTimeFormat):
            return jsonify(dict(err=1, msg="开始时间不能等于结束时间", data=[]))

        dataList = BEOPDataAccess.getInstance().getFixByPeriod(strTimeFrom, strTimeTo)

        if dataList is None:
            return jsonify(dict(err=1, msg="获取失败", data=[]))
        if not len(dataList):
            return jsonify(dict(err=1, msg="数据为空", data=[]))
        return jsonify(dict(err=0, msg="获取成功", data=dataList))

    except Exception as e:
        strLog = "ERROR in /fix/getByPeriod: %s" % e.__str__()
        logging.error(strLog)
        return jsonify(dict(err=1, msg="获取失败", data=[]))

@bp_fix.route("/keywordSearch", methods=["POST"])
def keyword_search():
    try:
        rcv = request.get_json()
        keyword = rcv.get("keyword") if rcv.get("keyword") is not None else None
        if keyword is None:
            return jsonify(dict(err=1, msg="关键词不能为空", data=[]))
        if not isinstance(keyword, str):
            return jsonify(dict(err=1, msg="关键词必须为字符串", data=[]))

        dataList = BEOPDataAccess.getInstance().keywordSearchFix(keyword)

        if dataList is None:
            return jsonify(dict(err=1, msg="获取失败", data=[]))
        if not len(dataList):
            return jsonify(dict(err=0, msg="无满足条件的记录", data=[]))
        return jsonify(dict(err=0, msg="获取成功", data=dataList))
    except Exception as e:
        strLog = "ERROR in /fix/keywordSearch: %s" % e.__str__()
        logging.error(strLog)
        return jsonify(dict(err=1, msg="搜索失败", data=[]))


@bp_fix.route("/downloadFixContentInPdf", methods=["POST"])
def download_fix_content_in_pdf():
    try:
        rcv = request.get_json()
        strTimeFormat = "%Y-%m-%d %H:%M:%S"
        levelDict = {0: "低", 1: "中", 2: "高"}
        resultDict = {0: "未解决", 1: "已解决"}

        fixId = rcv.get("fixId") if rcv.get("fixId") is not None else None
        if fixId is None:
            return jsonify(dict(err=1, msg="fixId不能为空", data=""))

        if not fixIdExists(fixId):
            return jsonify(dict(err=1, msg="该报修不存在", data=""))

        contentTuple = BEOPDataAccess.getInstance().getFixContentById(fixId)

        if not contentTuple:
            return jsonify(dict(err=1, msg="该报修无内容或获取内容失败", data=""))
        if not len(contentTuple):
            return jsonify(dict(err=1, msg="该报修无内容", data=""))

        reportTime = contentTuple[0].strftime(strTimeFormat)
        importance = levelDict.get(contentTuple[1], "低")
        urgent = levelDict.get(contentTuple[2], "低")
        strContent = contentTuple[3]
        result = resultDict.get(contentTuple[4], "无")
        closeTime = contentTuple[5].strftime(strTimeFormat) if contentTuple[5] > contentTuple[0] else "无"
        reportUser = contentTuple[6]
        solveUser = contentTuple[7] if len(contentTuple[7]) > 0 else "无"
        energyEffects = levelDict.get(contentTuple[8], "低")
        reportTitle = contentTuple[9]

        parList = process_fix_content(strContent)

        templateDocx = os.path.join(app.static_folder, "files", "fixReport.docx")
        if not os.path.exists(templateDocx):
            return jsonify(dict(err=1, msg="模板文件不存在，下载失败", data=""))
        if not is_zipfile(templateDocx):
            return jsonify(dict(err=1, msg="模板文件格式有误，下载失败", data=""))

        # 准备中间docx文件路径
        strNow = datetime.now().strftime("%Y%m%d%H%M%S")
        filesDir = os.path.join(app.static_folder, "files")
        if not os.path.exists(filesDir):
            os.mkdir(filesDir)

        tempDocxName = "tempDocx_{0}.docx".format(strNow)
        tempDocxPath = os.path.join(filesDir, tempDocxName)
        if os.path.exists(tempDocxPath):
            os.remove(tempDocxPath)

        bSucceed = genReportDocx(templateDocx, tempDocxPath, reportTitle, reportTime, reportUser, urgent, importance, energyEffects, solveUser, result, closeTime, parList)

        if not bSucceed:
            return jsonify(dict(err=1, msg="下载失败", data=""))

        destPdfName = "fixReportPdf_{0}_{1}.pdf".format(fixId, strNow)
        destPdfPath = os.path.join(filesDir, destPdfName)
        bSucceed = genReportPdf(tempDocxPath, destPdfPath)
        if not bSucceed:
            return jsonify(dict(err=1, msg="下载失败", data=""))

        os.remove(tempDocxPath)
        return jsonify(dict(err=0, msg="下载成功", data=destPdfName))

    except Exception as e:
        strLog = "ERROR in /fix/downloadFixContentInPdf: %s" % e.__str__()
        logging.error(strLog)
        return jsonify(dict(err=1, msg="下载失败", data=""))

@bp_fix.route("/getAll")
def get_all():
    try:
        itemList = BEOPDataAccess.getInstance().getFixContentByResult()
        dataList = []
        for item in itemList:
            dataList.append({
                "reportTime": item[0].strftime("%Y-%m-%d"),
                "content": htmlToStr(item[1]),
                "urgent": item[2],
                "importance": item[3],
                "energyEffects": item[4],
                "reportUser": item[5],
                "title": item[6],
                "result": item[7],
                "id": item[8],
                "closeTime": item[9].strftime("%Y-%m-%d")
            })
        return jsonify(dict(err=0, msg="获取成功", data=dataList))
    except Exception as e:
        strLog = "ERROR in /fix/getAll: %s" % e.__str__()
        logging.error(strLog)
        return jsonify(dict(err=1, msg="获取失败", data=[]))
