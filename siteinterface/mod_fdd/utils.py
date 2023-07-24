# -*- coding: utf-8 -*-
from siteinterface.BEOPDataAccess import BEOPDataAccess
from siteinterface.RedisManager import RedisManager
from datetime import datetime
import logging
from siteinterface.utils import g_work_order_inspect_time
from siteinterface import app
import os
import re
from openpyxl import Workbook
from bs4 import BeautifulSoup
from version import siteVersion
import math
from siteinterface.OSSTool import OSSTool

MaintainanceFaultPointName = "dom_system_maintainance_fault"
workOrderStatusMapZh = {0: "待分派", 1: "进行中", 2: "待确认", 3: "已确认", 4: "暂停", 5: "已终止"}
workOrderStatusMapEn = {0: "toBeAssigned", 1: "ongoing", 2: "toBeConfirmed", 3: "confirmed", 4: "pending", 5: "terminated"}

# 工单的评论图片在OSS的存放路径
OSS_FDD_IMG_DIR = "static/images/fdd/"

"""
opType: (操作类型)  -1:创建； 0: 修改状态；1: 修改预计完成时间
"""


def get_fault_label_point(faultPointName):
    if faultPointName.find("Fault") == -1:
        return ""

    faultLabelPoint = ""
    try:
        pointSplitList = faultPointName.split("Fault")
        faultLabelPoint = pointSplitList[0] + "FaultLabel" + pointSplitList[1]
        return faultLabelPoint
    except:
        return ""

def get_admin_of_executor(executor, authDict):
    if not len(executor):
        return ""

    executorGroupList = authDict.get(executor).get("group")
    for strName, dInfo in authDict.items():
        if dInfo.get("role") == "executor":
            continue

        for strGroup in executorGroupList:
            if strGroup in dInfo.get("group"):
                return strName
    return None

def time_to_inspect_work_order(tCurTime):
    tLastInspect = g_work_order_inspect_time
    tTar = tCurTime.replace(hour=3, minute=0, second=0, microsecond=0)
    if tLastInspect:
        if tLastInspect <= tTar <= tCurTime:
            return True
    return False

def inspect_work_order_on_time():
    global g_work_order_inspect_time

    tCurTime = datetime.now()
    if time_to_inspect_work_order(tCurTime):
        inspect_work_order()
    else:
        print("it's not time to inspect work order")
    g_work_order_inspect_time = tCurTime

"""
允许出现分派按钮
"""
def allow_assign(strRole, nStatus, strOwner, strUserName):
    if nStatus == 5:
        return 0
    if strRole == "admin":
        if nStatus in [4, 0]:
            return 1
        else:
            return 0

    elif strRole == "executor":
        if strOwner == strUserName:
            return 1
        else:
            return 0
    return 0

"""
允许出现提交按钮
"""
def allow_submit(strRole, nStatus, strOwner, strUserName, strCreator):
    if nStatus == 5:
        return 0

    if strRole == "admin":
        if nStatus in [0, 1]:
            return 0
        else:
            return 1

    if strRole == "executor":
        if nStatus == 1 and strOwner == strUserName:
            return 1
        elif nStatus == 0 and strUserName == strCreator:
            return 1
    return 0

"""
允许出现冻结按钮
"""
def allow_enable(strRole, fddName, MaintainanceFaultPointName):
    if fddName == MaintainanceFaultPointName:
        return 0
    if strRole == "admin":
        return 1
    else:
        return 0

"""
允许出现终止按钮
"""
def allow_terminate(nStatus, strRole):
    nAllow = 0
    if strRole == "admin":
        if nStatus in [0, 1, 2, 3, 4]:
            nAllow = 1
    return nAllow

"""
允许出现暂停按钮
nStatus: current fault status
"""
def allow_pause(nStatus, strRole):
    nAllow = 0
    if strRole == "admin" and nStatus in [1]:
        nAllow = 1
    return nAllow

"""
允许出现admin的专属提交键
"""
def allow_admin_submit(nStatus, strRole):
    nAllow = 0
    if strRole == "admin" and nStatus in [0, 1, 2, 4]:
        nAllow = 1
    return nAllow

"""
允许出现编辑按钮
"""
def allow_edit(nStatus, strRole):
    nAllow = 0
    if strRole == "admin" and nStatus in [0, 1, 4]:
        nAllow = 1
    return nAllow

"""
每天3:00检查一遍工单
"""
def inspect_work_order():
    try:
        authDict = RedisManager.get("FaultAuthInfo")
        resOngoingPendingOrders = BEOPDataAccess.getInstance().getAllOngoingAndPendingFaultOrders()
        if resOngoingPendingOrders.get("code") > 0:
            logging.error("UpdateWorkOrderError: 数据库查询失败")
            return dict(code=1, msg="数据库查询失败，请稍后再试", data=False)

        # ongoingPendingOrderrList: orderId, modifyTime, status, fddName, opUserName, opType, opContent, createTime, ownUser, detail, title
        ongoingPendingOrderrList = resOngoingPendingOrders.get("data", [])

        toResetList = []
        inValidList = []
        for order in ongoingPendingOrderrList:
            if order[8] not in authDict.keys():
                toResetList.append(
                    (order[3], "system", 0, order[0], datetime.now().strftime("%Y-%m-%d %H:%M:%S"), order[9], 0, order[7], order[10])
                )
                inValidList.append(
                    dict(orderId=order[0], owner=order[8], status=order[2])
                )

        if not len(toResetList):
            return dict(code=0, msg="检查完毕，未发现无效处理人", data=True)

        strInfo = ""
        for dInvalid in inValidList:
            strInfo += "(工单ID:{id},所属人:{name},状态:{status});".format(id=dInvalid.get("orderId"),
                                                                    name=dInvalid.get("owner"),
                                                                    status=workOrderStatusMapZh.get(dInvalid.get("status")))

        strLog = "共发现{count}个异常工单：{strInfo}".format(count=len(inValidList), strInfo=strInfo)
        logging.error("ResetWorkOrder: {log}".format(log=strLog))

        bSuc = BEOPDataAccess.getInstance().resetWorkOrders(toResetList)

        return dict(code=0 if bSuc else 1,
                    msg=strLog + ",成功设置为待分派" if bSuc else ",设置为待分派失败",
                    data=bSuc)

    except Exception as e:
        strLog = "UpdateWorkOrderError: {err}".format(err=e.__str__())
        logging.error(strLog)
        return dict(code=1, msg=strLog, data=False)

def has_chinese_character(strTar):
    for chara in strTar:
        if u"\u4e00" <= chara <= u"\u9fff":
            return True
    return False


def saveFddImgToLocal(fileList):
    try:
        staticFolder = app.static_folder
        fileDir = os.path.join(staticFolder, "files")
        if not os.path.exists(fileDir):
            os.mkdir(fileDir)

        fddImagsDir = os.path.join(fileDir, "fddImages")
        if not os.path.exists(fddImagsDir):
            os.mkdir(fddImagsDir)

        fileNameList = []
        for file in fileList:
            strFileName = "{0}_{1}".format(datetime.now().strftime("%Y%m%d%H%M%S"), file.filename)
            strFilePath = os.path.join(fddImagsDir, strFileName)
            if os.path.exists(strFilePath):
                os.remove(strFilePath)
            file.save(strFilePath)

            nFileSize = os.path.getsize(strFilePath)
            nFileSizeM = nFileSize / float(1024 * 1024)
            if nFileSizeM > 10:
                os.remove(strFilePath)
                return dict(err=1, msg="图片大小必须小于10M", data="")

            fileNameList.append(strFileName)

        return dict(err=0, msg="上传成功", data=fileNameList)

    except Exception as e:
        strLog = "ERROR in saveFddImgToLocal: %s" % e.__str__()
        logging.error(strLog)
        return dict(err=1, msg="上传失败", data="")

def is_p_html(strTar):
    try:
        rt = re.match(r"<p><img src=.*?><br></p>", strTar)
        if rt:
            return True
        return False
    except:
        return False

def insert_work_order_into_excel(infoList):
    titleList = ["工单号", "故障名称", "详情", "位置", "类型", "等级", "发生时间", "分组", "处理人", "状态", "结论", "开始处理时间", "预计完成时间", "处理时长"]
    try:
        book = Workbook()
        sheet = book.create_sheet("工单", 0)

        # 写表头
        for idx, item in enumerate(titleList):
            sheet.cell(1, idx+1, item)

        for idx, dInfo in enumerate(infoList):
            orderId = dInfo.get("id")
            orderName = dInfo.get("name")
            detail = dInfo.get("detail","无")
            position = dInfo.get("position", "无")
            cate = dInfo.get("view", "无")
            level = dInfo.get("level", "无")
            happenTime = dInfo.get("time", "无")
            group = dInfo.get("group", "无")
            processor = dInfo.get("processor", "无")
            status = dInfo.get("status", "无")
            conclusion = dInfo.get("conclusion", "无")
            startTime = dInfo.get("startTime", "无")
            estimatedTime = dInfo.get("estimatedTime", "无")
            duration = dInfo.get("duration")

            valueList = [orderId, orderName,detail,position,cate, level,happenTime,group,processor,status,conclusion,startTime,estimatedTime,duration]
            for i, value in enumerate(valueList):
                sheet.cell(row=idx+2, column=i+1, value=value)

        strNow = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")

        filesDir = os.path.join(app.static_folder, "files")
        if not os.path.exists(filesDir):
            os.mkdir(filesDir)

        tempDir = os.path.join(filesDir, "temp")
        if not os.path.exists(tempDir):
            os.mkdir(tempDir)

        destExcelName = "workOrder_{0}.xlsx".format(strNow)
        destExcelPath = os.path.join(tempDir, destExcelName)

        if os.path.exists(destExcelPath):
            os.remove(destExcelPath)

        book.save(destExcelPath)

        return destExcelName

    except Exception as e:
        logging.error("ERROR in insert_work_order_into_excel: %s" % e.__str__())
        return ""

def get_detail_from_p_html(strDetailWithHtml):
    try:
        if strDetailWithHtml.find("<p>") == -1:
            return strDetailWithHtml

        strText = ""
        soup = BeautifulSoup(strDetailWithHtml, "lxml")
        pTagList = soup.find_all("p")
        for pTag in pTagList:
            strText += pTag.get_text()
        return strText
    except:
        return strDetailWithHtml


def get_proc_version():
    dVersion = {}
    dCoreLogicVersion = BEOPDataAccess.getInstance().getCoreLogicVersion()
    strPysiteVersion = siteVersion.getCurrentVersion()
    dVersion.update(dict(dompysite=strPysiteVersion, domcore=dCoreLogicVersion.get("version", {}).get("domcore", "")))
    return dVersion

"""
按创建时间降列排序评论流
"""
def descending_order_comment_by_create_time(commentList):
    for i in range(len(commentList)):
        for j in range(len(commentList)-i-1):
            if commentList[j]["createTime"] < commentList[j+1]["createTime"]:
                commentList[j], commentList[j+1] = commentList[j+1], commentList[j]
    return commentList

"""
时间追溯描述
"""
def time_desc_trace_back(tPass):
    try:
        if not isinstance(tPass, datetime):
            return None

        if tPass > datetime.now():
            return None

        tDeltaSecs = (datetime.now() - tPass).total_seconds()
        if tDeltaSecs < 60:
            return "1分钟内"
        elif tDeltaSecs < 3600:
            return "{}分钟前".format(math.floor(tDeltaSecs / 60))
        elif tDeltaSecs < 3600 * 24:
            return "{}小时前".format(math.floor(tDeltaSecs / 3600))
        else:
            return "{}天前".format(math.floor(tDeltaSecs / (3600*24)))
    except:
        return None




