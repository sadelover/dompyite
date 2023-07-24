from siteinterface.mod_point import bp_point
from siteinterface import app
from flask import request, jsonify
from logic.StandardPointName import *
from .chunk import TABLE_HEAD, SOURCE_LIST
from openpyxl import Workbook
import os
from datetime import datetime
from .utils import *
import traceback
from siteinterface.BEOPSqliteAccess import BEOPSqliteAccess
import json
import requests
from siteinterface.BEOPDataAccess import BEOPDataAccess
from siteinterface.OSSTool import OSSTool
from siteinterface.mod_point.utils import get_source_type_list


@bp_point.route("/editEquationForVPoint", methods=["POST"])
def point_edit_equation_vpoint():
    try:
        rv = request.get_json()

        strPointName = rv.get("pointName")
        if strPointName is None or strPointName=='':
            return jsonify(dict(err=1, msg="pointName param error", data=""))

        strEquation = rv.get("equation")
        if strEquation is None:
            return jsonify(dict(err=1, msg="equation param error", data=""))

        strDescription = rv.get("description")
        if strDescription is None:
            return jsonify(dict(err=1, msg="description param error", data=""))

        bSuccess = BEOPSqliteAccess.getInstance().saveVPointCalculationEquation(strPointName, strEquation, strDescription)
        return jsonify(dict(err=0, msg='ok', data=dict(result=bSuccess)))
    except Exception as e:
        traceback.print_exc()
        return jsonify(dict(err=1, msg='ERROR in point_edit_equation_vpoint:' + e.__str__(), data=''))


@bp_point.route("/autoGeneratePointList", methods=["POST"])
def auto_generate_point_list():
    try:
        rv = request.get_json()
        res = dict()
        strPlantName = rv.get("strPlantName") if rv.get("strPlantName") is not None else ""
        strPlantDesc = rv.get("strPlantDesc") if rv.get("strPlantDesc") is not None else ""
        nChCount = rv.get("nCh") if rv.get("nCh") is not None else 0
        listCT = rv.get("listCT") if rv.get("listCT") is not None else list()
        nCWPCount = rv.get("nCWP") if rv.get("nCWP") is not None else 0
        nPriChWP = rv.get("nPriChWP") if rv.get("nPriChWP") is not None else 0
        nSecChWP = rv.get("nSecChWP") if rv.get("nSecChWP") is not None else 0
        nPriHWP = rv.get("nPriHWP") if rv.get("nPriHWP") is not None else 0
        nSecHWP = rv.get("nSecHWP") if rv.get("nSecHWP") is not None else 0

        bGenControl  = rv.get("bGenCtrlPoint") if rv.get("bGenCtrlPoint") is not None else 0
        if isinstance(bGenControl, int):
            bGenControl = (bGenControl==1)

        if not is_int([nChCount, nCWPCount, nPriChWP, nSecChWP, nPriHWP,
                       nSecHWP]):
            return jsonify(dict(err=1, msg="生成失败，设备数量必须为整数", data=""))
        if not isinstance(listCT, list):
            return jsonify(dict(err=1, msg="生成失败，冷却塔必须是一个列表"))
        if not is_str([strPlantName, strPlantDesc]):
            return jsonify(dict(err=1, msg="生成失败，点名和注释中的机房名称必须为字符串", data=""))


        if nChCount:
            listPoint = list()
            strEquip = "冷机"
            for idx in range(nChCount):
                for obj in g_standard_chiller:
                    listPoint.append({
                        "pointName": "%s%s%02d" % (strPlantName, obj.get("pointName"), (idx + 1)),
                        "description": "%s%s-%02d#%s" % (strPlantDesc, strEquip, (idx + 1), obj.get("description"))
                    })
            res.update({"chiller": listPoint})

        if nPriChWP:
            listPoint = list()
            strEquip = "一次冷冻泵"
            for idx in range(nPriChWP):
                for obj in g_standard_prichwp:
                    listPoint.append({
                        "pointName": "%s%s%02d" % (strPlantName, obj.get("pointName"), (idx + 1)),
                        "description": "%s%s-%02d#%s" % (strPlantDesc, strEquip, (idx + 1), obj.get("description"))
                    })
            res.update({"prichwp": listPoint})

        if nSecChWP:
            listPoint = list()
            strEquip = "二次冷冻泵"
            for idx in range(nSecChWP):
                for obj in g_standard_secchwp:
                    listPoint.append({
                        "pointName": "%s%s%02d" % (strPlantName, obj.get("pointName"), (idx + 1)),
                        "description": "%s%s-%02d#%s" % (strPlantDesc, strEquip, (idx + 1), obj.get("description"))
                    })
            res.update({"secchwp": listPoint})

        if nCWPCount:
            listPoint = list()
            strEquip = "冷却泵"
            for idx in range(nCWPCount):
                for obj in g_standard_cwp:
                    listPoint.append({
                        "pointName": "%s%s%02d" % (strPlantName, obj.get("pointName"), (idx + 1)),
                        "description": "%s%s-%02d#%s" % (strPlantDesc, strEquip, (idx + 1), obj.get("description"))
                    })
            res.update({"cwp": listPoint})

        if len(listCT):
            listPoint = list()
            strEquip = "冷却塔"
            for ctIdx, ct in enumerate(listCT):
                # 冷却塔
                for obj in g_standard_ct:
                    listPoint.append({
                        "pointName": "%s%s%02d" % (strPlantName, obj.get("pointName"), (ctIdx + 1)),
                        "description": "%s%02d#%s%s" % (strPlantDesc, (ctIdx + 1), strEquip, obj.get("description"))
                    })
                # 冷却塔风机
                for fanIdx in range(ct.get("nFan")):
                    for obj in g_standard_ctfan:
                        listPoint.append({
                            "pointName": "%sCTCTFan%02d%s%02d" % (
                            strPlantName, (fanIdx + 1), obj.get("pointName"), (ctIdx + 1)),
                            "description": "%s冷却塔-%02d#冷却塔风机%s%02d" % (
                            strPlantDesc, (fanIdx + 1), obj.get("pointName"), (ctIdx + 1))
                        })
            res.update({"ct": listPoint})

        if nPriHWP:
            listPoint = list()
            strEquip = "一次热水泵"
            for idx in range(nPriHWP):
                for obj in g_standard_prihwp:
                    listPoint.append({
                        "pointName": "%s%s%02d" % (strPlantName, obj.get("pointName"), (idx + 1)),
                        "description": "%s%s-%02d#%s" % (strPlantDesc, strEquip, (idx + 1), obj.get("description"))
                    })
            res.update({"prihwp": listPoint})

        if nSecHWP:
            listPoint = list()
            strEquip = "二次热水泵"
            for idx in range(nSecHWP):
                for obj in g_standard_sechwp:
                    listPoint.append({
                        "pointName": "%s%s%02d" % (strPlantName, obj.get("pointName"), (idx + 1)),
                        "description": "%s%s-%02d#%s" % (strPlantDesc, strEquip, (idx + 1), obj.get("description"))
                    })
            res.update({"sechwp": listPoint})

        # system common
        listPoint = list()
        for obj in g_standard_system_comm:
            listPoint.append({
                "pointName": "%s%s" % (strPlantName, obj.get("pointName")),
                "description": "%s%s" % (strPlantDesc, obj.get("description"))
            })
        res.update({"system_comm": listPoint})

        # system secondary
        if nSecChWP:
            listPoint = list()
            for obj in g_standard_system_sec:
                listPoint.append({
                    "pointName": "%s%s" % (strPlantName, obj.get("pointName")),
                    "description": "%s%s" % (strPlantDesc, obj.get("description"))
                })
            res.update({"system_sec": listPoint})

        # 优化控制
        if bGenControl:
            listPoint = list()
            for obj in g_standard_opt_ctrl:
                listPoint.append({
                    "pointName": "%s%s" % (strPlantName, obj.get("pointName")),
                    "description": "%s%s" % (strPlantDesc, obj.get("description"))
                })
            res.update({"opt_ctrl": listPoint})

        # 一键开关冷机
        if bGenControl and nChCount > 0:
            listPoint = list()
            for idx in range(nChCount):
                for obj in g_standard_opt_ctrl_chiller:
                    listPoint.append({
                        "pointName": "%s%s%02d" % (strPlantName, obj.get("pointName"), (idx + 1)),
                        "description": "%s%s%02d#冷机" % (strPlantDesc, obj.get("description"), (idx + 1))
                    })
            res.update({"opt_ctrl_chiller": listPoint})

        # 点表合并
        collect = list()
        for key in res.keys():
            category = res.get(key)
            collect += category

        book = Workbook()

        temp_folder = os.path.join(app.static_folder, "files", "temp")
        file_name = 'auto_gen_point_table_%s.xlsx'%(datetime.now().strftime('%Y%m%d%H%M%S'))
        saveFilePath = os.path.join(temp_folder, file_name)
        if os.path.exists(saveFilePath):
            os.remove(saveFilePath)

        sheet = book.create_sheet("Sheet1", 0)

        # 写表头
        for idx, item in enumerate(TABLE_HEAD):
            sheet.cell(row=1, column=(idx + 1), value=item)

        for idx, item in enumerate(collect):
            # 写序号
            sheet.cell(row=(idx+2), column=1, value=(idx+1))
            sheet.cell(row=(idx+2), column=2, value=item.get("pointName"))
            sheet.cell(row=(idx+2), column=3, value="vpoint")
            sheet.cell(row=(idx+2), column=4, value=item.get("description"))
            sheet.cell(row=(idx + 2), column=21, value=2)

        book.save(saveFilePath)

        return jsonify(dict(err=0, msg='点表生成成功', data= 'files/temp/'+ file_name))
    except:
        traceback.print_exc()
        return jsonify(dict(err=1, msg='点表生成失败', data=''))


@bp_point.route("/detectMissingModbusEquipmentPointsByLocalConfig")
def detect_missing_modbus_equipment_points_by_local_config():
    try:
        strModbusClientConfig = BEOPSqliteAccess.getInstance().getValueByKeyInLocalConfig("modbusclientconfig")
        if strModbusClientConfig == None:
            return jsonify(dict(err=1, msg="ModbusEquipment配置获取失败，可能未配置", data={}))

        headers = {"Content-Type": "application/json"}
        postData = {"jsonString": strModbusClientConfig}

        rcv = requests.post("http://47.100.17.99/api/pointtool/genModbusEquipmentPoints", data=json.dumps(postData), headers=headers, timeout=15)

        if not rcv:
            return jsonify(dict(err=1, msg="云端ModbusEquipment点名列表获取失败(无返回)", data={}))
        if rcv.status_code != 200:
            return jsonify(dict(err=1, msg="云端ModbusEquipment点名列表获取失败(status_code不等于200)", data={}))

        dRcv = json.loads(rcv.text)

        if dRcv.get("err", 1) == 1:
            return jsonify(dict(err=1, msg="云端ModbusEquipment点名列表生成失败", data={}))

        modbusEquipMsg = dRcv.get("msg", "")
        modbusEquipPointList = dRcv.get("data")

        curPointList = BEOPSqliteAccess.getInstance().getPointListForExport()

        curPointNameList = []
        for point in curPointList:
            curPointNameList.append(point[1].decode("gbk"))

        toInsertList = []
        toInsertPointNameList = []

        for modPointName in modbusEquipPointList:
            if modPointName.get("point", ""):
                if modPointName.get("point") not in curPointNameList:
                    strRW = "W" if modPointName.get("point").lower().find("setting") != -1 or modPointName.get("point").lower().find("setpoint") != -1 else "R"

                    if modPointName.get("point") in toInsertPointNameList:
                        continue

                    toInsertList.append(
                        dict(name=modPointName.get("point"),
                             type=modPointName.get("type"),
                             description=modPointName.get("description", ""),
                             rw=strRW,
                             unit=modPointName.get("unit", ""))
                    )
                    toInsertPointNameList.append(modPointName.get("point"))

        if not len(toInsertList):
            return jsonify(dict(err=0, msg="未发现需要增加的ModbusEquipment相关点", data={"info": modbusEquipMsg, "pointList": []}))
        return jsonify(dict(err=0, msg="获取成功", data={"info": modbusEquipMsg, "pointList": toInsertList}))

    except Exception as e:
        return jsonify(dict(err=1, msg="获取缺失ModbusEquipment相关点失败: %s" % e.__str__(), data={"info": "", "pointList": []}))


@bp_point.route("/insertModbusEquipmentPoints", methods=["POST"])
def insert_modbus_equipment_points():
    try:
        rcv = request.get_json()
        pointList = rcv.get("pointList", [])
        if not len(pointList):
            return jsonify(dict(err=1, msg="pointList不能为空", data={}))

        curPointList = BEOPSqliteAccess.getInstance().getPointListForExport()
        curMaxIndex = BEOPSqliteAccess.getInstance().getMaxIdInTable("list_point")
        nNum = curMaxIndex + 1

        curPointNameList = []
        for curPoint in curPointList:
            curPointNameList.append(curPoint[1].decode("gbk"))

        toInsertList = []
        toInsertPointNameList = []
        for point in pointList:
            name = point.get("name", None)
            strType = point.get("type", None)
            if name == None or strType == None:
                continue

            if name in curPointNameList:
                continue

            if name in toInsertPointNameList:
                continue

            # id, name, type, description, unit, rw, param1, param2, param3, param4, param5, param6, param7, param8, param9, param10, high, highhigh, low, lowlow, param11(存储周期), param15, param16, param17, param18, param19
            toInsertList.append(
                [str(nNum), name, strType, point.get("description", ""), point.get("unit", ""), point.get("rw", "R"),
                 '', '', '', '', '', '', '', '', '', '', '0', '0', '0', '0', '2', '', '', '', '', "2"]
            )

            toInsertPointNameList.append(name)

            nNum += 1

        if not len(toInsertList):
            return jsonify(dict(err=0, msg="未发现需要增加的ModbusEquipment相关点，或传入的点信息格式有误", data={"addCount": 0}))

        dResult = BEOPSqliteAccess.getInstance().insertPointList(toInsertList)
        if dResult.get("success", False):
            BEOPSqliteAccess.getInstance().updateAllPointInfoIntoRedis()
            return jsonify(dict(err=0, msg="新增ModbusEquipment相关点成功", data={"addCount": len(toInsertList)}))
        else:
            return jsonify(dict(err=1, msg="新增ModbusEquipment相关点失败:{msg}".format(msg=dResult.get("msg","")), data={"addCount": 0}))

    except Exception as e:
        return jsonify(dict(err=1, msg="新增ModbusEquipment相关点失败: %s" % e.__str__(), data={"addCount": 0}))


@bp_point.route("/addOnePoint", methods=["POST"])
def add_one_point():
    try:
        rcv = request.get_json()
        name = rcv.get("name", None)
        source = rcv.get("sourceType", None)
        description = rcv.get("description", "")
        unit = rcv.get("unit", "")
        rw = rcv.get("RW", "0")
        param1 = rcv.get("addr", "")
        param2 = rcv.get("param2", "")
        param3 = rcv.get("param3", "")
        param4 = rcv.get("param4", "")
        param5 = rcv.get("param5", "")
        param6 = rcv.get("param6", "")
        param7 = rcv.get("param7", "")
        param8 = rcv.get("param8", "")
        param9 = rcv.get("param9", "")
        param10 = rcv.get("param10", "")
        high = rcv.get("high", "0")
        highhigh = rcv.get("highhigh", "0")
        low = rcv.get("low", "0")
        lowlow = rcv.get("lowlow", "0")
        storageCycle = rcv.get("storecycle", "2")
        custom = rcv.get("customName", "")
        system = rcv.get("system", "")
        equipment = rcv.get("device", "")
        sysType = rcv.get("type", "")
        decimal = rcv.get("decimal", "2")

        if name == None:
            return jsonify(dict(err=1, msg="点名不能为空", data=False))
        if not isinstance(name, str):
            return jsonify(dict(err=1, msg="点名必须为字符串", data=False))
        if not len(name):
            return jsonify(dict(err=1, msg="点名不能为空", data=False))
        if re.match(r"^[0-9]{1}$", name[0]):
            return jsonify(dict(err=1, msg="点名首字母不能为数字", data=False))
        if name.find(" ") >= 0:
            return jsonify(dict(err=1, msg="点名不能含有空格", data=False))
        if not re.match(r"^[a-zA-Z0-9_]*$", name):
            return jsonify(dict(err=1, msg="点名不能含有特殊字符", data=False))

        if BEOPSqliteAccess.getInstance().pointExists(name):
            return jsonify(dict(err=1, msg="该点名已存在", data=False))

        if source == None:
            source = "vpoint"

        sourceInfoList = get_source_type_list()
        srcList = [source["name"] for source in sourceInfoList]
        if source not in srcList:
            return jsonify(dict(err=1, msg="未知source", data=False))

        if not isinstance(description, str):
            return jsonify(dict(err=1, msg="注释必须为字符串", data=False))
        if not isinstance(unit, str):
            return jsonify(dict(err=1, msg="单位必须为字符串", data=False))

        nRW = None
        try:
            nRW = int(rw)
        except:
            pass
        if nRW not in [0, 1]:
            return jsonify(dict(err=1, msg="读写属性必须为0或1", data=False))

        if not isinstance(param1, str):
            return jsonify(dict(err=1, msg="param1必须为字符串", data=False))
        if not isinstance(param2, str):
            return jsonify(dict(err=1, msg="param2必须为字符串", data=False))
        if not isinstance(param3, str):
            return jsonify(dict(err=1, msg="param3必须为字符串", data=False))
        if not isinstance(param4, str):
            return jsonify(dict(err=1, msg="param4必须为字符串", data=False))
        if not isinstance(param5, str):
            return jsonify(dict(err=1, msg="param5必须为字符串", data=False))
        if not isinstance(param6, str):
            return jsonify(dict(err=1, msg="param6必须为字符串", data=False))
        if not isinstance(param7, str):
            return jsonify(dict(err=1, msg="param7必须为字符串", data=False))
        if not isinstance(param8, str):
            return jsonify(dict(err=1, msg="param8必须为字符串", data=False))
        if not isinstance(param9, str):
            return jsonify(dict(err=1, msg="param9必须为字符串", data=False))
        if not isinstance(param10, str):
            return jsonify(dict(err=1, msg="param10必须为字符串", data=False))
        if not isinstance(high, str):
            return jsonify(dict(err=1, msg="high必须为字符串", data=False))
        if not is_digit(high):
            return jsonify(dict(err=1, msg="high必须为数字字符串", data=False))

        if not isinstance(highhigh, str) and not isinstance(highhigh, int):
            return jsonify(dict(err=1, msg="highhigh必须为字符串或整数", data=False))

        if not isinstance(low, str) and not isinstance(low, int):
            return jsonify(dict(err=1, msg="low必须为字符串或整数", data=False))

        if not isinstance(lowlow, str):
            return jsonify(dict(err=1, msg="lowlow必须为字符串", data=False))
        if not is_digit(lowlow):
            return jsonify(dict(err=1, msg="lowlow必须为数字字符串", data=False))

        if not isinstance(storageCycle, str):
            return jsonify(dict(err=1, msg="存储周期必须为数字字符串", data=False))
        if storageCycle not in ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9"]:
            return jsonify(dict(err=1, msg="存储周期必须为0-9的整数", data=False))
        if not isinstance(custom, str):
            return jsonify(dict(err=1, msg="custom必须为字符串", data=False))
        if not isinstance(system, str):
            return jsonify(dict(err=1, msg="system必须为字符串", data=False))
        if not isinstance(equipment, str):
            return jsonify(dict(err=1, msg="equipment必须为字符串", data=False))

        if not isinstance(sysType, str):
            return jsonify(dict(err=1, msg="设备类型必须为字符串", data=False))
        if not isinstance(decimal, str):
            return jsonify(dict(err=1, msg="小数位必须为字符串", data=False))
        if not is_digit(decimal):
            return jsonify(dict(err=1, msg="小数位必须为整数字符串", data=False))

        nMaxId = BEOPSqliteAccess.getInstance().getMaxIdInTable("list_point")

        # id, name, type, description, unit, rw,
        # param1, param2, param3, param4, param5, param6, param7, param8, param9, param10, high, highhigh, low, lowlow, param11(存储周期), 自定义, 系统, 设备, 类型, 小数位
        toInsertList = []
        toInsertList.append(
            [str(nMaxId+1), name, source, description, unit, str(nRW),
             param1, param2, param3, param4.encode("gbk"), param5, param6, param7, param8, param9, param10, high, str(highhigh), str(low), lowlow, storageCycle, custom, system, equipment, sysType, decimal]
        )

        dResult = BEOPSqliteAccess.getInstance().insertPointList(toInsertList)

        if dResult.get("success", False) == True and source == "vpoint":
            BEOPDataAccess.getInstance().operateUnit01ForVpointAddDel(name, 1)

        BEOPSqliteAccess.getInstance().updateAllPointInfoIntoRedis()

        if dResult.get("success", False):
            return jsonify(dict(err=0, msg="点%s添加成功" % name, data=True))
        else:
            return jsonify(dict(err=1, msg="点%s添加失败" % name, data=False))

    except Exception as e:
        strLog = "新增点失败:%s" % (e.__str__())
        return jsonify(dict(err=1, msg=strLog, data=False))

@bp_point.route("/deletePoint", methods=["POST"])
def delete_point():
    rcv = request.get_json()
    point = rcv.get("point", [])

    if isinstance(point, str):
        pointList = [point]
    elif isinstance(point, list):
        pointList = point
    else:
        return jsonify(dict(err=1, msg="point必须为字符串（单点）或列表（多点）", data=False))

    dPointInfo = BEOPSqliteAccess.getInstance().getPointInfoFromS3db(pointList)
    toDeleteVpointNameList = []
    for pName in pointList:
        if dPointInfo.get(pName, {}).get("sourceType", "") == "vpoint":
            toDeleteVpointNameList.append(pName)

    if len(toDeleteVpointNameList):
        BEOPDataAccess.getInstance().operateUnit01ForVpointAddDel(toDeleteVpointNameList, 0)

    result = BEOPSqliteAccess.getInstance().deletePoint(pointList)
    BEOPSqliteAccess.getInstance().updateAllPointInfoIntoRedis()
    if result.get("success", False):
        return jsonify(dict(err=0, msg="删除成功", data=True))
    return jsonify(dict(err=1, msg="删除失败", data=False))

@bp_point.route("/updatePoint", methods=["POST"])
def update_point():
    rcv = request.get_json()
    id = rcv.get("id", None)
    name = rcv.get("name", None)
    source = rcv.get("sourceType", None)
    description = rcv.get("description", None)
    unit = rcv.get("unit", None)
    rw = rcv.get("RW", None)
    param1 = rcv.get("addr", None)
    param2 = rcv.get("param2", None)
    param3 = rcv.get("param3", None)
    param4 = rcv.get("param4", None)
    param5 = rcv.get("param5", None)
    param6 = rcv.get("param6", None)
    param7 = rcv.get("param7", None)
    param8 = rcv.get("param8", None)
    param9 = rcv.get("param9", None)
    param10 = rcv.get("param10", None)
    high = rcv.get("high", None)
    highhigh = rcv.get("highhigh", None)
    low = rcv.get("low", None)
    lowlow = rcv.get("lowlow", None)
    storageCycle = rcv.get("storecycle", None)
    custom = rcv.get("customName", None)
    system = rcv.get("system", None)
    equipment = rcv.get("device", None)
    sysType = rcv.get("type", None)
    decimal = rcv.get("decimal", None)

    if id == None:
        return jsonify(dict(err=1, msg="id不能为空", data=False))

    if not is_digit(id):
        return jsonify(dict(err=1, msg="id必须为数字", data=False))

    nId = int(id)

    if name != None:
        if not isinstance(name, str):
            return jsonify(dict(err=1, msg="点名必须为字符串", data=False))

        bDuplicated = BEOPSqliteAccess.getInstance().hasDuplicatedPointName(name, id)
        if bDuplicated == None:
            return jsonify(dict(err=1, msg="点名重名检查失败", data=False))

        if bDuplicated == True:
            return jsonify(dict(err=1, msg="%s在现有点表中重名" % name, data=False))

    if source != None:
        sourceInfoList = get_source_type_list()
        srcList = [source["name"] for source in sourceInfoList]
        if source not in srcList:
            return jsonify(dict(err=1, msg="未知点位类型", data=False))

    if description != None:
        if not isinstance(description, str):
            return jsonify(dict(err=1, msg="注释必须为字符串", data=False))

    if unit != None:
        if not isinstance(unit, str):
            return jsonify(dict(err=1, msg="单位必须为字符串", data=False))

    nRW = None
    if rw != None:
        try:
            nRW = int(rw)
        except:
            pass
        if nRW not in [0, 1]:
            return jsonify(dict(err=1, msg="读写属性必须为0或1", data=False))

    if param1 != None:
        if not isinstance(param1, str):
            return jsonify(dict(err=1, msg="param1必须为字符串", data=False))
    if param2 != None:
        if not isinstance(param2, str):
            return jsonify(dict(err=1, msg="param2必须为字符串", data=False))
    if param3 != None:
        if not isinstance(param3, str):
            return jsonify(dict(err=1, msg="param3必须为字符串", data=False))
    if param4 != None:
        if not isinstance(param4, str):
            return jsonify(dict(err=1, msg="param4必须为字符串", data=False))
    if param5 != None:
        if not isinstance(param5, str):
            return jsonify(dict(err=1, msg="param5必须为字符串", data=False))
    if param6 != None:
        if not isinstance(param6, str):
            return jsonify(dict(err=1, msg="param6必须为字符串", data=False))
    if param7 != None:
        if not isinstance(param7, str):
            return jsonify(dict(err=1, msg="param7必须为字符串", data=False))
    if param8 != None:
        if not isinstance(param8, str):
            return jsonify(dict(err=1, msg="param8必须为字符串", data=False))
    if param9 != None:
        if not isinstance(param9, str):
            return jsonify(dict(err=1, msg="param9必须为字符串", data=False))
    if param10 != None:
        if not isinstance(param10, str):
            return jsonify(dict(err=1, msg="param10必须为字符串", data=False))
    if high != None:
        if not isinstance(high, str):
            return jsonify(dict(err=1, msg="high必须为字符串", data=False))
        if not is_digit(high):
            return jsonify(dict(err=1, msg="high必须为数字字符串", data=False))

    if highhigh != None:
        if not isinstance(highhigh, str) and not isinstance(highhigh, int):
            return jsonify(dict(err=1, msg="highhigh必须为字符串或整数", data=False))

    if low != None:
        if not isinstance(low, str) and not isinstance(low, int):
            return jsonify(dict(err=1, msg="low必须为字符串或整数", data=False))

    if lowlow != None:
        if not isinstance(lowlow, str):
            return jsonify(dict(err=1, msg="lowlow必须为字符串", data=False))
        if not is_digit(lowlow):
            return jsonify(dict(err=1, msg="lowlow必须为数字字符串", data=False))

    if storageCycle != None:
        if not isinstance(storageCycle, str):
            return jsonify(dict(err=1, msg="存储周期必须为数字字符串", data=False))
        try:
            storageCycle = int(float(storageCycle))
            storageCycle = str(storageCycle)
        except:
            return jsonify(dict(err=1, msg="存储周期必须为0-9的整数", data=False))

        if storageCycle not in ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9"]:
            return jsonify(dict(err=1, msg="存储周期必须为0-9的整数", data=False))

    if custom != None:
        if not isinstance(custom, str):
            return jsonify(dict(err=1, msg="custom必须为字符串", data=False))
    if system != None:
        if not isinstance(system, str):
            return jsonify(dict(err=1, msg="system必须为字符串", data=False))
    if equipment != None:
        if not isinstance(equipment, str):
            return jsonify(dict(err=1, msg="equipment必须为字符串", data=False))

    if sysType != None:
        if not isinstance(sysType, str):
            return jsonify(dict(err=1, msg="设备类型必须为字符串", data=False))
    if decimal != None:
        if not isinstance(decimal, str):
            return jsonify(dict(err=1, msg="小数位必须为字符串", data=False))
        if not is_digit(decimal):
            return jsonify(dict(err=1, msg="小数位必须为整数字符串", data=False))

    bSuc = BEOPSqliteAccess.getInstance().updatePoint(nId, name, source, description, unit, nRW, param1, param2, param3, param4, param5, param6, param7, param8,
                                               param9, param10, high, str(highhigh), str(low), lowlow, storageCycle, custom, system, equipment, sysType, decimal)
    BEOPSqliteAccess.getInstance().updateAllPointInfoIntoRedis()
    return jsonify(dict(err=1 if not bSuc else 0, msg="", data=bSuc))

"""
根据输入的中文注释从StandardPointDescMap.json中匹配中文获取点名和注释
"""
@bp_point.route("/matchStandardPoint", methods=["POST"])
def match_standard_point():
    rcv = request.get_json()
    if not isinstance(rcv, dict):
        rcv = {}

    descInput = rcv.get("desc", None)
    if not isinstance(descInput, str):
        return jsonify(dict(err=1, msg="传入的中文注释不能为空", data=False))
    if not len(descInput):
        return jsonify(dict(err=1, msg="传入的中文注释不能为空", data=False))

    tempDir = os.path.join(app.static_folder, "temp")
    if not os.path.exists(tempDir):
        os.mkdir(tempDir)

    lfPath = os.path.join(tempDir, "StandardPointDescMap_{time}.json".format(time=datetime.now().strftime("%Y-%m-%d-%H-%M-%S-%f")))
    if os.path.exists(lfPath):
        try:
            os.remove(lfPath)
        except:
            pass

    tool = OSSTool()
    bSuc = tool.download("update/dbwizard/StandardPointDescMap.json", lfPath)
    if not bSuc or not os.path.exists(lfPath):
        try:
            os.remove(lfPath)
        except:
            pass
        return jsonify(dict(err=1, msg="从云端获取标准点名库失败，请稍后重试", data={}))

    with open(lfPath, "r", encoding="UTF8", errors="ignore") as fo:
        try:
            dJson = json.load(fo)
        except:
            dJson = {}

    if not isinstance(dJson, dict):
        try:
            os.remove(lfPath)
        except:
            pass
        return jsonify(dict(err=1, msg="从云端获取到的标准点名库格式有误，请稍后重试", data={}))

    dAllMap = dJson.get("data", {})
    fileIntroduction = dJson.get("introduction", None)
    if not dAllMap or not isinstance(dAllMap, dict):
        try:
            os.remove(lfPath)
        except:
            pass
        return jsonify(dict(err=1, msg="从云端获取标准点名库失败，请稍后重试", data={}))

    stdPointNameDescMapLocalFilePath = os.path.join(tempDir, "ChillerPlantEquipStandardPointNameAndDesc_{time}.json".format(time=datetime.now().strftime("%Y-%m-%d-%H-%M-%S-%f")))
    if os.path.exists(stdPointNameDescMapLocalFilePath):
        try:
            os.remove(stdPointNameDescMapLocalFilePath)
        except:
            pass

    dMatch = {}
    similarList = []
    recommendList = []
    for desc, dInfo in dAllMap.items():
        if desc == descInput and dInfo:
            dMatch = dict(point=dInfo.get("point", ""),
                           description=desc,
                           intro=dInfo.get("intro", ""))

        if descInput.find(desc) >= 0 and desc != descInput:
            similarList.append(dict(point=dInfo.get("point", ""),
                                    description=desc,
                                    intro=dInfo.get("intro", "")))

    # 未匹配到
    if not dMatch:
        if descInput not in dAllMap.keys():
            dAllMap.update({descInput: {}})

        if dAllMap:
            if not fileIntroduction:
                fileIntroduction = "标准点名注释映射表"
            dJson = {"data": dAllMap, "introduction": fileIntroduction}

            with open(lfPath, "w", encoding="UTF8", errors="ignore") as fo:
                try:
                    fo.write(str(dJson).replace("'", '"'))
                except:
                    try:
                        os.remove(lfPath)
                    except:
                        pass
                    return jsonify(dict(err=1, msg="操作失败", data={}))

            tool.upload(lfPath, "update/dbwizard/StandardPointDescMap.json")

        # 若未匹配到则从ChillerPlantEquipStandardPointNameAndDesc.json中根据注释查询一次
        bSuc = tool.download("update/dbwizard/ChillerPlantEquipStandardPointNameAndDesc.json",
                             stdPointNameDescMapLocalFilePath)
        if bSuc and os.path.exists(stdPointNameDescMapLocalFilePath):
            with open(stdPointNameDescMapLocalFilePath, "r", encoding="UTF8", errors="ignore") as fo:
                try:
                    dJson = json.load(fo)
                except:
                    dJson = {}

            dCommonEquipmentPointDescMap = dJson.get("data", {})
            for strType, dType in dCommonEquipmentPointDescMap.items():  # 设备、传感器、执行器
                for strEquipName, dEquip in dType.items():  # 冷机、水泵
                    equipAttr = dEquip.get("equipAttr", "")
                    equipAttrCh = dEquip.get("equipAttrCh", "")
                    dPointAttr = dEquip.get("pointAttr", {})
                    for strPoint, strDesc in dPointAttr.items():
                        if strDesc == descInput:
                            recommendList.append(dict(point=strPoint,
                                                      description=strDesc,
                                                      intro="点名举例：{equipAttr}{strPoint}01，注释：1#{equipAttrCh}{strDesc}".format(
                                                          equipAttr=equipAttr,
                                                          strPoint=strPoint,
                                                          equipAttrCh=equipAttrCh,
                                                          strDesc=strDesc)))

    try:
        os.remove(lfPath)
    except:
        pass

    try:
        os.remove(stdPointNameDescMapLocalFilePath)
    except:
        pass

    strMsg = "匹配成功"
    if not dMatch and not len(recommendList):
        strMsg = "未匹配到注释为'{descInput}'的标准点名，已在云端记录".format(descInput=descInput)

    return jsonify(dict(err=0, msg=strMsg, data={"match": dMatch, "similarList": similarList, "recommendList": recommendList}))

@bp_point.route("/getSourceType")
def get_source_type():
    srcList = get_source_type_list()
    return jsonify(dict(err=0, msg="", data=srcList))

"""
获取常见设备的高频标准点名及释义列表
"""
@bp_point.route("/getStandardPointNameAndDesc")
def get_standard_point_name_and_desc():
    tempDir = os.path.join(app.static_folder, "temp")
    if not os.path.exists(tempDir):
        os.mkdir(tempDir)

    fileName = "ChillerPlantEquipStandardPointNameAndDesc_{time}.json".format(time=datetime.now().strftime("%Y-%m-%d-%H-%M-%S-%f"))
    filePath = os.path.join(tempDir, fileName)
    if os.path.exists(filePath):
        try:
            os.remove(filePath)
        except:
            return jsonify(dict(err=1, msg="删除上次残留文件失败，请稍后再试", data=[]))

    tool = OSSTool()
    bSuc = tool.download("update/dbwizard/ChillerPlantEquipStandardPointNameAndDesc.json", filePath)
    if not bSuc or not os.path.exists(filePath):
        return jsonify(dict(err=1, msg="json文件下载失败", data=[]))

    with open(filePath, "r", encoding="UTF8", errors="ignore") as fo:
        try:
            dJson = json.load(fo)
        except:
            try:
                os.remove(filePath)
            except:
                pass
            return jsonify(dict(err=1, msg="云端JSON配置有误，解析失败", data=[]))

    if not isinstance(dJson, dict):
        try:
            os.remove(filePath)
        except:
            pass
        return jsonify(dict(err=1, msg="云端JSON配置有误", data=[]))

    dAllData = dJson.get("data", [])
    if not isinstance(dAllData, dict):
        try:
            os.remove(filePath)
        except:
            pass
        return jsonify(dict(err=1, msg="云端JSON配置有误", data=[]))

    strTypeList = list(dAllData.keys())   # 设备、传感器、执行器
    for i in range(len(strTypeList)):
        for j in range(len(strTypeList)-i-1):
            if strTypeList[j] > strTypeList[j+1]:
                strTypeList[j], strTypeList[j+1] = strTypeList[j+1], strTypeList[j]

    res = []
    for strType in strTypeList:
        dEquip = dAllData.get(strType, {})
        equipNameList = list(dEquip.keys())
        for i in range(len(equipNameList)):
            for j in range(len(equipNameList) - i - 1):
                if equipNameList[j] > equipNameList[j + 1]:
                    equipNameList[j], equipNameList[j + 1] = equipNameList[j + 1], equipNameList[j]

        facilityList = []
        for equipName in equipNameList:  # 冷机、水泵
            dEquipDetail = dEquip.get(equipName, {})
            equipAttr = dEquipDetail.get("equipAttr", "")
            equipAttrCh = dEquipDetail.get("equipAttrCh", "")
            dPointAttr = dEquipDetail.get("pointAttr", {})
            pList = []
            for pName, desc in dPointAttr.items():
                pList.append(dict(point=pName, desc=desc))

            for i in range(len(pList)):
                for j in range(len(pList)-i-1):
                    if pList[j]["point"] > pList[j+1]["point"]:
                        pList[j], pList[j+1] = pList[j+1], pList[j]

            facilityList.append(dict(name=equipName, children=pList))
        res.append(dict(name=strType, children=facilityList))

    try:
        os.remove(filePath)
    except:
        pass
    return jsonify(dict(err=0, msg="获取成功", data=res))


@bp_point.route("/addPoints", methods=["POST"])
def add_points():
    try:
        rcv = request.get_json()
        pointList = rcv.get("pointList", None)

        if not isinstance(pointList, list):
            return jsonify(dict(err=1, msg="pointList必须是数组", data=False))

        sourceInfoList = get_source_type_list()
        srcList = [source["name"] for source in sourceInfoList]

        nMaxId = BEOPSqliteAccess.getInstance().getMaxIdInTable("list_point")

        bAllVpoint = True

        nPointIndex = nMaxId + 1
        toInsertList = []
        toInsertPointNameList = []
        for dPoint in pointList:
            name = dPoint.get("name", None)
            source = dPoint.get("sourceType", None)
            description = dPoint.get("description", "")
            unit = dPoint.get("unit", "")
            rw = dPoint.get("RW", "0")
            param1 = dPoint.get("addr", "")
            param2 = dPoint.get("param2", "")
            param3 = dPoint.get("param3", "")
            param4 = dPoint.get("param4", "")
            param5 = dPoint.get("param5", "")
            param6 = dPoint.get("param6", "")
            param7 = dPoint.get("param7", "")
            param8 = dPoint.get("param8", "")
            param9 = dPoint.get("param9", "")
            param10 = dPoint.get("param10", "")
            high = dPoint.get("high", "0")
            highhigh = dPoint.get("highhigh", "0")
            low = dPoint.get("low", "0")
            lowlow = dPoint.get("lowlow", "0")
            storageCycle = dPoint.get("storecycle", "2")
            custom = dPoint.get("customName", "")
            system = dPoint.get("system", "")
            equipment = dPoint.get("device", "")
            sysType = dPoint.get("type", "")
            decimal = dPoint.get("decimal", "2")

            if name == None:
                continue
            if not isinstance(name, str):
                continue
            if not len(name):
                continue
            if re.match(r"^[0-9]{1}$", name[0]):
                continue
            if name.find(" ") >= 0:
                continue
            if not re.match(r"^[a-zA-Z0-9_]*$", name):
                continue

            if BEOPSqliteAccess.getInstance().pointExists(name):
                continue

            if source == None:
                source = "vpoint"

            if source not in srcList:
                continue

            if source != "vpoint":
                bAllVpoint = False

            if not isinstance(description, str):
                continue
            if not isinstance(unit, str):
                continue

            nRW = None
            try:
                nRW = int(rw)
            except:
                pass
            if nRW not in [0, 1]:
                continue

            if not isinstance(param1, str):
                continue
            if not isinstance(param2, str):
                continue
            if not isinstance(param3, str):
                continue
            if not isinstance(param4, str):
                continue
            if not isinstance(param5, str):
                continue
            if not isinstance(param6, str):
                continue
            if not isinstance(param7, str):
                continue
            if not isinstance(param8, str):
                continue
            if not isinstance(param9, str):
                continue
            if not isinstance(param10, str):
                continue
            if not isinstance(high, str):
                continue
            if not is_digit(high):
                continue

            if not isinstance(highhigh, str) and not isinstance(highhigh, int):
                continue

            if not isinstance(low, str) and not isinstance(low, int):
                continue

            if not isinstance(lowlow, str):
                continue
            if not is_digit(lowlow):
                continue

            if not isinstance(storageCycle, str):
                continue
            if storageCycle not in ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9"]:
                continue
            if not isinstance(custom, str):
                continue
            if not isinstance(system, str):
                continue
            if not isinstance(equipment, str):
                continue

            if not isinstance(sysType, str):
                continue
            if not isinstance(decimal, str):
                continue
            if not is_digit(decimal):
                continue

            # id, name, type, description, unit, rw,
            # param1, param2, param3, param4, param5, param6, param7, param8, param9, param10, high, highhigh, low, lowlow, param11(存储周期), 自定义, 系统, 设备, 类型, 小数位

            toInsertList.append(
                [str(nPointIndex), name, source, description, unit, str(nRW),
                 param1, param2, param3, param4, param5, param6, param7, param8, param9, param10, high, str(highhigh), str(low), lowlow, storageCycle, custom, system, equipment, sysType, decimal]
            )

            toInsertPointNameList.append(name)

            nPointIndex += 1

        if not len(toInsertList):
            return jsonify(dict(err=1, msg="未发现合法的点列表可执行插入", data=False))

        dResult = BEOPSqliteAccess.getInstance().insertPointList(toInsertList)

        if dResult.get("success", False) == True and bAllVpoint == True:
            BEOPDataAccess.getInstance().operateUnit01ForVpointAddDel(toInsertPointNameList, 1)

        BEOPSqliteAccess.getInstance().updateAllPointInfoIntoRedis()

        if dResult.get("success", False):
            return jsonify(dict(err=0, msg="成功添加{n}个点".format(n=len(toInsertList)), data=True))
        else:
            return jsonify(dict(err=1, msg="添加失败", data=False))

    except Exception as e:
        strLog = "新增多个点失败:%s" % (e.__str__())
        return jsonify(dict(err=1, msg=strLog, data=False))


