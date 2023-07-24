from siteinterface.mod_env import bp_env
from siteinterface import app
from flask import request, jsonify
from logic.StandardPointName import *
from openpyxl import Workbook
import os
from datetime import datetime
import traceback
from .. BEOPMySqlDBContainer import BEOPMySqlDBContainer
from siteinterface.BEOPSqliteAccess import *
from siteinterface.BEOPDataAccess import BEOPDataAccess
import logging
from .utils import *
import xlrd


@bp_env.route("/create", methods=["POST"])
def create():
    try:
        rcv = request.get_json()
        ena = rcv.get("enabled") if rcv.get("enabled") is not None else 1
        name = rcv.get("name") if rcv.get("name") else None
        description = rcv.get("description") if rcv.get("description") else ""
        tags = rcv.get("tags") if rcv.get("tags") else ""
        creator = rcv.get("creator") if rcv.get("creator") else None
        createtime = datetime.now()

        # if type:
        #     if not isinstance(type, int):
        #         return jsonify(dict(err=1, msg="type 必须是一个整数或空字符串"))

        if ena:
            if not isinstance(ena, int):
                return jsonify(dict(err=1, msg="enabled 必须是一个整数或者不传入（不传入默认为1）"))

        if not name:
            return jsonify(dict(err=1, msg="无场景名称"))

        if not isinstance(name, str):
            return jsonify(dict(err=1, msg="场景名称必须是一个字符串"))

        # 检查场景名称是否已存在
        envNameExists = BEOPDataAccess.getInstance().checkIfEnvNameExists(name)
        if envNameExists["code"] > 0:
            return jsonify(dict(err=1, msg="场景名检查失败"))
        if envNameExists["data"]:
            return jsonify(dict(err=1, msg="该场景名称已被占用，请换一个"))

        if not isinstance(description, str):
            return jsonify(dict(err=1, msg="场景描述必须是一个字符串"))

        if not isinstance(tags, str):
            return jsonify(dict(err=1, msg="场景标签必须是一个字符串"))

        if not creator:
            return jsonify(dict(err=1, msg="无场景创建者"))

        if not isinstance(creator, str):
            return jsonify(dict(err=1, msg="场景创建者必须是一个字符串"))

        maxId = BEOPDataAccess.getInstance().getMaxIdInTable("env")

        envid = maxId + 1

        # tag处理
        strTag = None
        if tags == "":
            strTag = ""
        elif tags is not None:
            tagList = []
            for tag in tags.split(";"):
                strTar = tag.strip()
                if len(strTar) and strTar not in tagList:
                    tagList.append(strTar)

            strTag = ";".join(tagList)

        rvConduct = BEOPDataAccess.getInstance().createEnv(envid, ena, name, description, strTag, creator, createtime)

        if rvConduct:
            return jsonify(dict(err=0, msg="创建成功"))
        else:
            return jsonify(dict(err=1, msg="创建失败"))

    except Exception as e:
        strLog = "ERROR in /env/create: %s" % e.__str__()
        logging.error(strLog)

        return jsonify(dict(err=1, msg="创建失败: %s" % e.__str__()))

@bp_env.route("/remove", methods=["POST"])
def remove():
    try:
        rcv = request.get_json()
        envid = rcv.get("id") if rcv.get("id") else None

        if not envid:
            return jsonify(dict(err=1, msg="场景ID不能为空"))

        if not isinstance(envid, int):
            return jsonify(dict(err=1, msg="场景ID必须是一个整数"))

        if not envIdExistsInEnv(envid):
            return jsonify(dict(err=1, msg="该场景ID不存在"))

        strSource = rcv.get("source", "")
        strIP = request.remote_addr
        strLog = '[IMPORTANT]removeEnv: 来自IP:%s, 用户:%s, envId:%s' % (strIP, strSource, envid)
        log_info_to_file("dompysite_mode_env_{date}.log".format(date=datetime.now().strftime("%Y-%m-%d")), strLog)

        rvConduct = BEOPDataAccess.getInstance().removeEnv(envid)

        if rvConduct:
            return jsonify(dict(err=0, msg="删除成功"))
        else:
            return jsonify(dict(err=1, msg="删除失败"))

    except Exception as e:
        strLog = "ERROR in /env/remove: %s" % e.__str__()
        logging.error(strLog)

        return jsonify(dict(err=1, msg="删除失败: %s" % e.__str__()))

@bp_env.route("/edit", methods=["POST"])
def edit():
    try:
        rcv = request.get_json()

        envid = rcv.get("id") if rcv.get("id") is not None else None
        type = rcv.get("type") if rcv.get("type") is not None else None
        enabled = rcv.get("enabled") if rcv.get("enabled") is not None else None
        name = rcv.get("name") if rcv.get("name") is not None else None
        description = rcv.get("description") if rcv.get("description") is not None else None
        tags = rcv.get("tags") if rcv.get("tags") is not None else None
        creator = rcv.get("creator") if rcv.get("creator") is not None else None

        if envid is None:
            return jsonify(dict(err=1, msg="场景ID不能为空"))

        if not isinstance(envid, int):
            return jsonify(dict(err=1, msg="场景ID必须是一个整数"))

        if not envIdExistsInEnv(envid):
            return jsonify(dict(err=1, msg="该场景不存在"))

        if type:
            if not isinstance(type, int):
                return jsonify(dict(err=1, msg="type 必须是一个整数或空字符串"))

        if enabled:
            if not isinstance(enabled, int):
                return jsonify(dict(err=1, msg="enabled 必须是一个整数或者不传入（不传入默认为1）"))

        if name:
            if not isinstance(name, str):
                return jsonify(dict(err=1, msg="场景名称必须是一个字符串"))

        # 检查场景名称是否已存在
        nameUsed = BEOPDataAccess.getInstance().nameUsedByOtherEnv(name, envid)
        if nameUsed["code"] > 0:
            return jsonify(dict(err=1, msg="检查场景名称失败"))
        else:
            if nameUsed["data"]:
                return jsonify(dict(err=1, msg="该场景名称已被使用，请换一个"))

        if description:
            if not isinstance(description, str):
                return jsonify(dict(err=1, msg="场景描述必须是一个字符串"))

        if tags:
            if not isinstance(tags, str):
                return jsonify(dict(err=1, msg="场景标签必须是一个字符串"))

        if creator:
            if not isinstance(creator, str):
                return jsonify(dict(err=1, msg="场景创建者必须是一个字符串"))

        # tag处理
        strTag = None
        if tags == "":
            strTag = ""
        elif tags is not None:
            tagList = []
            for tag in tags.split(";"):
                strTar = tag.strip()
                if len(strTar) and strTar not in tagList:
                    tagList.append(strTar)

            strTag = ";".join(tagList)

        paramList = []
        strSetList = []
        for item in [(type, "type"), (enabled, "enabled"), (name, "name"), (description, "description"), (strTag, "tags"), (creator, "creator")]:
            if item[0] is not None:
                paramList.append(item[0])
                strSetList.append("{0}=%s".format(item[1]))

        if not len(paramList):
            return jsonify(dict(err=0, msg="无内容需要更新"))

        strIP = request.remote_addr
        strSource = rcv.get("source", "")
        strLog = '[IMPORTANT]editEnv: 来自IP:%s, 用户:%s, envId:%s' % (strIP, strSource, envid)
        log_info_to_file("dompysite_mode_env_{date}.log".format(date=datetime.now().strftime("%Y-%m-%d")), strLog)

        bSuc = BEOPDataAccess.getInstance().editEnv(paramList, strSetList, envid)

        if bSuc:
            return jsonify(dict(err=0, msg="更新成功"))
        else:
            return jsonify(dict(err=1, msg="更新失败"))

    except Exception as e:
        strLog = "ERROR in /env/edit: %s" % e.__str__()
        logging.error(strLog)

        return jsonify(dict(err=1, msg="更新失败: %s" % e.__str__()))

@bp_env.route("/getAll", methods=["GET"])
def get_all():
    try:
        rvConduct = BEOPDataAccess.getInstance().getAllEnv()
        if rvConduct is None:
            return jsonify(dict(err=1, msg="数据库繁忙，请重试", data=None))

        res = list()
        if len(rvConduct):
            for item in rvConduct:
                res.append({
                    "id": item[0],
                    "type": item[1],
                    "enabled": item[2],
                    "name": item[3],
                    "description": item[4],
                    "tags": item[5],
                    "creator": item[6],
                    "createtime": item[7].strftime("%Y-%m-%d %H:%M:%S")
                })

        return jsonify(dict(err=0, msg="获取成功", data=res))

    except Exception as e:
        strLog = "ERROR in /env/getAll: %s" % e.__str__()
        logging.error(strLog)
        return jsonify(dict(err=1, msg="获取失败", data=list()))

@bp_env.route("/get", methods=["POST"])
def get_by_id():
    try:
        rcv = request.get_json()
        envid = rcv.get("id") if rcv.get("id") is not None else None

        if not envid:
            return jsonify(dict(err=1, msg="场景ID不能为空"))

        if not isinstance(envid, int):
            return jsonify(dict(err=1, msg="场景ID必须是一个整数"))

        if not BEOPDataAccess.getInstance().envIdExistsInEnv(envid):
            return jsonify(dict(err=1, msg="不存在ID为{0}的记录".format(envid), data={}))

        dEnv, dEnvDetail = BEOPDataAccess.getInstance().getEnvById(envid)
        if dEnv == None or dEnvDetail == None:
            return jsonify(dict(err=1, msg="数据库繁忙，请重试", data={}))

        dData = {
            "id": dEnv[envid].get("id"),
            "type": dEnv[envid].get("type"),
            "enabled": dEnv[envid].get("enabled"),
            "name": dEnv[envid].get("name"),
            "description": dEnv[envid].get("description"),
            "tags": dEnv[envid].get("tags"),
            "creator": dEnv[envid].get("creator"),
            "createtime": dEnv[envid].get("createtime").strftime("%Y-%m-%d %H:%M:%S"),
            "detail": []
        }

        if len(dEnvDetail.get(envid, [])):

            pointNameList = [obj[0] for obj in dEnvDetail.get(envid)]
            pointValueList = [obj[1] for obj in dEnvDetail.get(envid)]

            pointInfoDict = BEOPSqliteAccess.getInstance().getPointInfoFromS3db_WithRedis(pointNameList)

            for idx, pointName in enumerate(pointNameList):
                try:
                    if pointInfoDict.get(pointName):
                        desc = pointInfoDict.get(pointName).get("description")
                        dData["detail"].append({
                            "pointName": pointName,
                            "pointValue": pointValueList[idx],
                            "description": desc
                        })
                except:
                    pass

            # dData["detail"].append({
            #     "pointName": [obj[0] for obj in arrDetail],
            #     "pointValue": [obj[1] for obj in arrDetail]
            # })

        return jsonify(dict(err=0, msg="获取成功", data=dData))

    except Exception as e:
        strLog = "ERROR in /env/get: %s" % e.__str__()
        logging.error(strLog)

        return jsonify(dict(err=1, msg="获取失败", data=dict()))


@bp_env.route("/import/<int:envid>", methods=["POST"])
def import_env_detail(envid=0):
    try:
        infoTable = request.files.get("file")

        if envid == 0:
            return jsonify(dict(err=1, msg="场景ID不能为0"))

        if not envIdExistsInEnv(envid):
            return jsonify(dict(err=1, msg="该场景ID不存在"))

        if not infoTable:
            strLog = "没有发现导入的表格"
            return jsonify(dict(err=1, msg=strLog))

        tableName = infoTable.filename
        if not tableName.endswith("xlsx"):
            return jsonify(dict(err=1, msg="只支持.xlsx文件"))

        saveFilePath = fetchEnvImportTable(infoTable, tableName)
        if not saveFilePath:
            return jsonify(dict(err=1, msg="获取导入的表格失败"))

        book = xlrd.open_workbook(saveFilePath)
        sheet = book.sheet_by_index(0)
        nrows = sheet.nrows

        if nrows <= 1:
            return jsonify(dict(err=1, msg="表格中无内容"))

        valueList = list()
        for idx in range(1, nrows):
            values = sheet.row_values(idx)
            valueList.append(values)

        pointNameListFromTable = [item[0] if len(item) else None for item in valueList]

        pointValueListFromTable = [item[1] if len(item) and len(item) > 1 else None for item in valueList]

        qualified = valuesFromTableQualified(valueList)
        if not qualified:
            return jsonify(dict(err=1, msg="导入的表格中信息配置有误", data=[]))

        if len(list(set(pointNameListFromTable))) != len(pointNameListFromTable):
            return jsonify(dict(err=1, msg="表格中点名重复", data=[]))

        # 导入表格后直接覆盖则不再检查数据库中是否已存在该点名
        # noDuplicate = noDuplicatePointsInEnvDetail(pointNameListFromTable, envid)
        # if not noDuplicate:
        #     return jsonify(dict(err=1, msg="相同的点名已存在于该场景ID下"))

        # 获取点位信息
        pointInfoDict = BEOPSqliteAccess.getInstance().getPointInfoFromS3db(pointNameListFromTable)

        if not pointInfoDict:
            return jsonify(dict(err=1, msg="导入失败，因导入的点表中没有一个点属于该项目", data=[]))

        pointNameListFrom4db = list(pointInfoDict.keys())

        invalidPointNameList = []
        validPointNameList = []
        for pointName in pointNameListFromTable:
            if pointName not in pointNameListFrom4db:
                invalidPointNameList.append(pointName)
                continue
            validPointNameList.append(pointName)

        rDataList = list()
        for idx, pointName in enumerate(validPointNameList):
            desc = pointInfoDict.get(pointName).get("description")
            rDataList.append({
                "pointName": pointName,
                "pointValue": pointValueListFromTable[idx],
                "description": desc
            })

        if len(invalidPointNameList):
            return jsonify(dict(err=2, msg="成功导入部分点位，因表格中存在无效点名", data=rDataList, invalid=invalidPointNameList))
        return jsonify(dict(err=0, msg="导入成功", data=rDataList))

    except Exception as e:
        strLog = "ERROR in /env/import: %s" % e.__str__()
        logging.error(strLog)

        return jsonify(dict(err=1, msg="导入失败：%s" % e.__str__(), data=[]))

#保存场景接口
@bp_env.route("/saveContent", methods=["POST"])
def save_content():
    try:
        rcv = request.get_json()
        envid = rcv.get("id") if rcv.get("id") else None
        pointNameList = rcv.get("pointNameList") if rcv.get("pointNameList") else list()
        pointValueList = rcv.get("pointValueList") if rcv.get("pointValueList") else list()

        if not envid:
            return jsonify(dict(err=1, msg="无场景ID"))

        # if not len(pointNameList):
        #     return jsonify(dict(err=1, msg="点名列表为空"))
        #
        # if not len(pointValueList):
        #     return jsonify(dict(err=1, msg="点值列表为空"))

        if len(pointNameList) != len(pointValueList):
            return jsonify(dict(err=1, msg="点名列表和点值列表的长度不一致"))

        if len(list(set(pointNameList))) != len(pointNameList):
            return jsonify(dict(err=1, msg="点名重复"))

        if not envIdExistsInEnv(envid):
            return jsonify(dict(err=1, msg="该场景ID不存在"))

        strIP = request.remote_addr
        strSource = rcv.get("source", "")
        strLog = '[IMPORTANT]saveContent: 来自IP:%s, 用户:%s, envId:%s' % (strIP, strSource, envid)
        log_info_to_file("dompysite_mode_env_{date}.log".format(date=datetime.now().strftime("%Y-%m-%d")), strLog)

        envList = list()
        for idx, pointName in enumerate(pointNameList):
            env = envid, pointName, pointValueList[idx]
            envList.append(env)

        bSuc = BEOPDataAccess.getInstance().updateEnvDetail(envid, envList)
        if bSuc:
            return jsonify(dict(err=0, msg="保存成功"))
        else:
            return jsonify(dict(err=1, msg="保存失败"))

    except Exception as e:
        strLog = "ERROR in /env/saveContent: %s" % e.__str__()
        logging.error(strLog)

        return jsonify(dict(err=1, msg="保存失败"))

"""
====  日程接口 ====
"""

@bp_env.route("/addSchedule", methods=["POST"])
def add_schedule():
    try:
        rcv = request.get_json()
        name = rcv.get("name") if rcv.get("name") else None
        envid = rcv.get("envId") if rcv.get("envId") else None    # 原point修改为envid
        author = rcv.get("author") if rcv.get("author") else None

        if not name:
            return jsonify(dict(err=1, msg="日程名称不能为空"))

        if not envid:
            return jsonify(dict(err=1, msg="场景ID不能为空"))

        if not author:
            return jsonify(dict(err=1, msg="创建者名称不能为空"))

        if not isinstance(envid, int):
            return jsonify(dict(err=1, msg="场景ID必须为整数"))

        if not isinstance(name, str) or not isinstance(author, str):
            return jsonify(dict(err=1, msg="日程名称和创建者名称必须为字符串"))

        if not envIdExistsInEnv(envid):
            return jsonify(dict(err=1, msg="该场景不存在"))

        strIP = request.remote_addr
        strSource = rcv.get("source", "")
        strLog = '[IMPORTANT]addSchedule: 来自IP:%s, 用户:%s, envId:%s, name:%s, author:%s' % (strIP, strSource, envid, name, author)
        log_info_to_file("dompysite_mode_env_{date}.log".format(date=datetime.now().strftime("%Y-%m-%d")), strLog)

        rvConduct = BEOPDataAccess.getInstance().addEnvSchedule(name, envid, author)

        if rvConduct:
            return jsonify(dict(err=0, msg="创建成功"))
        else:
            return jsonify(dict(err=1, msg="创建失败"))

    except Exception as e:
        strLog = "ERROR in /env/addSchedule: %s" % e.__str__()
        logging.error(strLog)
        return jsonify(dict(err=1, msg="创建失败"))

@bp_env.route("/removeSchedule", methods=["POST"])
def remove_schedule():
    try:
        rcv = request.get_json()

        scheduleId = rcv.get("scheduleId") if rcv.get("scheduleId") else None

        if not scheduleId:
            return jsonify(dict(err=1, msg="日程ID不能为空"))

        if not isinstance(scheduleId, int):
            return jsonify(dict(err=1, msg="日程ID必须为整数"))

        exists = BEOPDataAccess.getInstance().envScheduleExists(scheduleId)

        if not exists:
            return jsonify(dict(err=1, msg="该场景日程不存在"))

        strIP = request.remote_addr
        strSource = rcv.get("source", "")
        strLog = '[IMPORTANT]removeSchedule: 来自IP:%s, 用户:%s, scheduleId:%s' % (strIP, strSource, scheduleId)
        log_info_to_file("dompysite_mode_env_{date}.log".format(date=datetime.now().strftime("%Y-%m-%d")), strLog)

        rvConduct = BEOPDataAccess.getInstance().removeSchedule(scheduleId)

        if rvConduct:
            return jsonify(dict(err=0, msg="删除成功"))
        else:
            return jsonify(dict(err=1, msg="删除失败"))

    except Exception as e:
        strLog = "ERROR in /env/removeSchedule: %s" % e.__str__()
        logging.error(strLog)
        return jsonify(dict(err=1, msg="删除失败: %s" % e.__str__()))

@bp_env.route("/editSchedule", methods=["POST"])
def edit_schedule():
    try:
        rcv = request.get_json()
        scheduleId = rcv.get("scheduleId") if rcv.get("scheduleId") else None
        name = rcv.get("name") if rcv.get("name") else None
        envid = rcv.get("envId") if rcv.get("envId") else None
        author = rcv.get("author") if rcv.get("author") else None

        if not scheduleId:
            return jsonify(dict(err=1, msg="日程ID不能为空"))

        if not isinstance(scheduleId, int):
            return jsonify(dict(err=1, msg="日程ID必须为整数"))

        if not name:
            return jsonify(dict(err=1, msg="日程名称不能为空"))

        if not envid:
            return jsonify(dict(err=1, msg="场景ID不能为空"))

        if not isinstance(envid, int):
            return jsonify(dict(err=1, msg="场景ID必须为整数"))

        if not author:
            return jsonify(dict(err=1, msg="创建者名称不能为空"))

        exists = BEOPDataAccess.getInstance().envScheduleExists(scheduleId)
        if not exists:
            return jsonify(dict(err=1, msg="该日程不存在"))

        if not envIdExistsInEnv(envid):
            return jsonify(dict(err=1, msg="该场景不存在"))

        strIP = request.remote_addr
        strSource = rcv.get("source", "")
        strLog = '[IMPORTANT]removeSchedule: 来自IP:%s, 用户:%s, scheduleId:%s, name:%s, envId:%s, author:%s' % (strIP, strSource, scheduleId, name, envid, author)
        log_info_to_file("dompysite_mode_env_{date}.log".format(date=datetime.now().strftime("%Y-%m-%d")), strLog)

        rvConduct = BEOPDataAccess.getInstance().editEnvSchedule(name, envid, author, scheduleId)

        if rvConduct:
            return jsonify(dict(err=0, msg="更新成功"))
        else:
            return jsonify(dict(err=1, msg="更新失败"))

    except Exception as e:
        strLog = "ERROR in /env/editSchedule: %s" % e.__str__()
        logging.error(strLog)
        return jsonify(dict(err=1, msg="更新失败: %s" % e.__str__()))

@bp_env.route("/getAllSchedule", methods=["GET"])
def get_all_schedule():
    try:
        rvConduct = BEOPDataAccess.getInstance().getAllSchedule()
        if rvConduct is None:
            return jsonify(dict(err=1, msg="数据库繁忙，请重试", data=None))

        rvList = list()
        if len(rvConduct):
            for item in rvConduct:
                envId = convertStrIntoInt(item[3])
                rvList.append({
                    "scheduleId": item[0],
                    "name": item[2],
                    "envId": envId,
                    "loop": item[4],
                    "enable": item[5],
                    "author": item[6]
                })

        return jsonify(dict(err=0, msg="获取成功", data=rvList))

    except Exception as e:
        strLog = "ERROR in /env/editSchedule: %s" % e.__str__()
        logging.error(strLog)
        return jsonify(dict(err=1, msg="获取失败: %s" % e.__str__(), data=list()))

@bp_env.route("/getScheduleById", methods=["POST"])
def get_schedule_by_id():
    try:
        rcv = request.get_json()

        scheduleId = rcv.get("scheduleId") if rcv.get("scheduleId") else None

        if not isinstance(scheduleId, int):
            return jsonify(dict(err=1, msg="日程ID必须为整数"))

        exists = BEOPDataAccess.getInstance().envScheduleExists(scheduleId)

        if not exists:
            return jsonify(dict(err=1, msg="该场景日程不存在"))

        rvConduct = BEOPDataAccess.getInstance().getScheduleById(scheduleId)

        if rvConduct is None:
            return jsonify(dict(err=1, msg="获取失败", data={}))

        rvObj = dict()
        if len(rvConduct):
            rvObj = {
                "scheduleId": rvConduct[0][0],
                "name": rvConduct[0][2],
                "envId": convertStrIntoInt(rvConduct[0][3]),
                "loop": rvConduct[0][4],
                "enable": rvConduct[0][5],
                "author": rvConduct[0][6]
            }

        return jsonify(dict(err=0, msg="获取成功", data=rvObj))

    except Exception as e:
        strLog = "ERROR in /env/getScheduleById: %s" % e.__str__()
        logging.error(strLog)
        return jsonify(dict(err=1, msg="获取失败: %s" % e.__str__(), data=list()))

@bp_env.route("/getAllSchedulePlan", methods=["GET"])
def get_all_schedule_plan():
    try:
        rvConduct = BEOPDataAccess.getInstance().getAllEnvSchedulePlan()
        if rvConduct is None:
            return jsonify(dict(err=1, msg="数据库繁忙，请重试", data=None))


        planDict = {1:[], 2:[], 3:[], 4:[], 5:[], 6:[], 7:[]}

        # 将计划按星期几筛选归类
        for item in rvConduct:
            if item.get("weekday") in list(planDict.keys()):
                planDict.get(item.get("weekday")).append({
                    "planId": item.get("planId"),
                    "scheduleId": item.get("scheduleId"),
                    "time": item.get("time"),
                    "scheduleName": item.get("scheduleName")
                })

        rvPlan = list()
        for weekday in list(planDict.keys()):
            plan = planDict.get(weekday)

            # 将计划按  时间：计划 键值对做成一个字典
            weekdayPlanDict = dict()
            for item in plan:
                weekdayPlanDict.update({
                    item.get("time"): item
                })

            # 获取当日的计划时间列表
            weekdayTimeList = [datetime.strptime(item.get("time"), "%H:%S") for item in plan]

            # 冒泡降序处理当日的计划时间列表
            weekdayTimeListSorted = bubbleSort(weekdayTimeList)

            # 按照降序排列的时间列表排列当日计划
            weekdayPlanListSorted = list()
            for tTime in weekdayTimeListSorted:
                weekdayPlanListSorted.append(weekdayPlanDict.get(tTime.strftime("%H:%S")))

            # 装载
            rvPlan.append({
                "weekday": weekday,
                "planList": weekdayPlanListSorted
            })

        return jsonify(dict(err=0, msg="获取成功", data=rvPlan))

    except Exception as e:
        strLog = "ERROR in /env/getAllPlan: %s" % e.__str__()
        logging.error(strLog)
        return jsonify(dict(err=1, msg="获取失败: %s" % e.__str__(), data=list()))


@bp_env.route("/keywordSearch", methods=["POST"])
def keyword_search_env():
    try:
        rcv = request.get_json()
        keyword = rcv.get("keyword")

        rvConduct = BEOPDataAccess.getInstance().searchEnvByName(keyword)

        if rvConduct is None:
            return jsonify(dict(err=1, msg="数据库繁忙，请重试"),data=[])

        rvData = list()
        for item in rvConduct:
            rvData.append({
                "id": item[0],
                "type": item[1],
                "enable": item[2],
                "name": item[3],
                "description": item[4],
                "tags": item[5],
                "creator": item[6],
                "createtime": item[7].strftime("%Y-%m-%d %H:%M:%S")
            })

        return jsonify(dict(err=0, msg="获取成功", data=rvData))

    except Exception as e:
        strLog = "ERROR in /env/keywordSearch: %s" % e.__str__()
        logging.error(strLog)
        return jsonify(dict(err=1, msg="获取失败: %s" % e.__str__(), data=list()))

@bp_env.route("/loadSimulation", methods=["POST"])
def load_simulation():
    try:
        rcv = request.get_json()
        envId = rcv.get("envId", None)

        if envId is None:
            return jsonify(dict(err=1, msg="场景ID不能为空"))

        if not envIdExistsInEnv(envId):
            return jsonify(dict(err=1, msg="该场景不存在"))

        mode = BEOPDataAccess.getInstance().getCurrentMode()

        if mode is None:
            return jsonify(dict(err=1, msg="获取当前模式失败，无法加载仿真环境"))

        if mode != 0:
            return jsonify(dict(err=1, msg="系统处于非仿真模式，加载失败"))

        dEnv, dEnvDetail = BEOPDataAccess.getInstance().getEnvById(envId)
        if not dEnv or not dEnvDetail:
            return jsonify(dict(err=1, msg="数据库繁忙，请重试"))

        envDetailList = dEnvDetail.get(envId)
        if not envDetailList:
            return jsonify(dict(err=1, msg="未发现场景{id}的点点名-点值列表".format(id=environ)))

        pointNameList = []
        pointValueList = []
        for item in envDetailList:
            pointNameList.append(item[0])
            pointValueList.append(item[1])

        pointInfoDict = BEOPSqliteAccess.getInstance().getPointInfoFromS3db(pointNameList)

        if not pointInfoDict:
            return jsonify(dict(err=0, msg="场景中的点名不存在于该项目中"))

        pointNameListFrom4db = list(pointInfoDict.keys())

        for pointName in pointNameList:
            if pointName not in pointNameListFrom4db:
                return jsonify(dict(err=0, msg="场景中含有不存在于该项目的点名"))

        bSuc = BEOPDataAccess.getInstance().setRealtimeData(pointNameList, pointValueList)

        if bSuc.get("err") == 0:
            return jsonify(dict(err=0, msg="仿真环境加载成功"))
        else:
            return jsonify(dict(err=1, msg="仿真环境加载失败"))

    except Exception as e:
        strLog = "ERROR in /env/loadSimulation: %s" % e.__str__()
        logging.error(strLog)
        return jsonify(dict(err=1, msg="仿真环境加载失败: %s" % e.__str__()))

@bp_env.route("/saveSchedulePlan", methods=["POST"])
def save_schedule_plan():
    try:
        rcv = request.get_json()
        weekdayList = rcv.get("weekdayList") if rcv.get("weekdayList") else None
        strTime = rcv.get("time") if rcv.get("time") else None
        envIdList = rcv.get("envIdList") if rcv.get("envIdList") else None
        strScheduleName = rcv.get("scheduleName") if rcv.get("scheduleName") else None
        nLoop = rcv.get("loop") if rcv.get("loop") is not None else 1
        strAuthor = rcv.get("author") if rcv.get("author") else None

        if not weekdayList:
            return jsonify(dict(err=1, msg="周列表不能为空"))

        if not isinstance(weekdayList, list):
            return jsonify(dict(err=1, msg="周列表必须是一个数组"))

        if not strTime:
            return jsonify(dict(err=1, msg="时间不能为空"))

        if not strTime:
            return jsonify(dict(err=1, msg="时间不能为空"))

        if not isinstance(strTime, str):
            return jsonify(dict(err=1, msg="时间必须为一个字符串"))

        if not isValidDate(strTime, "%H:%S"):
            return jsonify(dict(err=1, msg="时间格式有误（时间格式示例：08:00）"))

        if not envIdList:
            return jsonify(dict(err=1, msg="场景ID列表不能为空"))

        if not isinstance(envIdList, list):
            return jsonify(dict(err=1, msg="场景ID列表必须是一个数组"))

        for envId in envIdList:
            if not envIdExistsInEnv(envId):
                return jsonify(dict(err=1, msg="有的场景ID不存在"))

        if not strScheduleName:
            return jsonify(dict(err=1, msg="日程名称不能为空"))

        if not isinstance(strScheduleName, str):
            return jsonify(dict(err=1, msg="日程名称必须为字符串"))

        if not isinstance(nLoop, int):
            return jsonify(dict(err=1, msg="周重复必须为整数"))

        if not strAuthor:
            return jsonify(dict(err=1, msg="创建者名称不能为空"))

        if not isinstance(strAuthor, str):
            return jsonify(dict(err=1, msg="创建者名称必须为字符串"))

        scheduleList = list()
        for envId in envIdList:
            scheduleList.append((1, strScheduleName, str(envId), nLoop, 1, strAuthor))

        planList = list()
        for weekday in weekdayList:
            for envId in envIdList:
                planList.append((weekday, strTime, strTime, 1, envId, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))

        strIP = request.remote_addr
        strSource = rcv.get("source", "")
        strLog = '[IMPORTANT]saveSchedulePlan: 来自IP:%s, 用户:%s, strScheduleName:%s' % (strIP, strSource, strScheduleName)
        log_info_to_file("dompysite_mode_env_{date}.log".format(date=datetime.now().strftime("%Y-%m-%d")), strLog)

        # 保存场景日程  表：schedule_list
        bSucSchedule = BEOPDataAccess.getInstance().insertEnvSchedule(scheduleList)

        # 保存场景日程计划  表：schedule_info_weeky
        bSucPlan = BEOPDataAccess.getInstance().insertEnvSchedulePlan(planList)

        if bSucSchedule and bSucPlan:
            return jsonify(dict(err=0, msg="创建成功"))
        else:
            return jsonify(dict(err=1, msg="创建失败"))

    except Exception as e:
        strLog = "ERROR in /env/saveSchedulePlan: %s" % e.__str__()
        logging.error(strLog)
        return jsonify(dict(err=1, msg="创建失败: %s" % e.__str__()))

@bp_env.route("/removeSchedulePlan", methods=["POST"])
def remove_schedule_plan():
    try:
        rcv = request.get_json()
        planId = rcv.get("planId") if rcv.get("planId") is not None else None

        if not planId:
            return jsonify(dict(err=1, msg="计划ID不能为空"))

        if not isinstance(planId, int):
            return jsonify(dict(err=1, msg="计划ID必须为整数"))

        strIP = request.remote_addr
        strSource = rcv.get("source", "")
        strLog = '[IMPORTANT]removeSchedulePlan: 来自IP:%s, 用户:%s, planId:%s' % (strIP, strSource, planId)
        log_info_to_file("dompysite_mode_env_{date}.log".format(date=datetime.now().strftime("%Y-%m-%d")), strLog)

        bSuc = BEOPDataAccess.getInstance().removeEnvSchedulePlan(planId)

        if bSuc:
            return jsonify(dict(err=0, msg="删除成功"))
        else:
            return jsonify(dict(err=1, msg="删除失败"))

    except Exception as e:
        strLog = "ERROR in /env/removeSchedulePlan: %s" % e.__str__()
        logging.error(strLog)
        return jsonify(dict(err=1, msg="删除失败: %s" % e.__str__()))


@bp_env.route("/export", methods=["POST"])
def export():
    try:
        rcv = request.get_json()
        envId = rcv.get("id") if rcv.get("id") else None

        if not envId:
            return jsonify(dict(err=1, msg="场景ID不能为空"))

        if not isinstance(envId, int):
            return jsonify(dict(err=1, msg="场景ID必须是一个整数"))

        if not envIdExistsInEnv(envId):
            return jsonify(dict(err=1, msg="该场景ID不存在"))

        dEnv, dEnvDetail = BEOPDataAccess.getInstance().getEnvById(envId)
        if not dEnv or not dEnvDetail:
            return jsonify(dict(err=1, msg="数据库繁忙，请重试"))

        if not len(dEnvDetail.get(envId, [])):
            return jsonify(dict(err=0, msg="无内容", data=""))

        detailList = dEnvDetail.get(envId)

        book = Workbook()
        sheet = book.create_sheet("场景详情", 0)

        # 写表头
        sheet.cell(row=1, column=1, value="pointname")
        sheet.cell(row=1, column=2, value="pointvalue")

        # 写表格内容
        for idx, detail in enumerate(detailList):
            sheet.cell(row=idx+2, column=1, value=detail[0])
            sheet.cell(row=idx+2, column=2, value=detail[1])

        strNow = datetime.now().strftime("%Y%m%d%H%M%S")

        filesDir = os.path.join(app.static_folder, "files")
        if not os.path.exists(filesDir):
            os.mkdir(filesDir)

        destExcelName = "envDetail_{0}.xlsx".format(strNow)
        destExcelPath = os.path.join(filesDir, destExcelName)

        if os.path.exists(destExcelPath):
            os.remove(destExcelPath)

        book.save(destExcelPath)

        return jsonify(dict(err=0, msg="下载成功", data=destExcelName))

    except Exception as e:
        strLog = "ERROR in /evn/export: %s" % e.__str__()
        logging.error(strLog)
        return jsonify(dict(err=1, msg="下载失败", data=""))

@bp_env.route("/getAllTags")
def get_all_tags_from_env():
    tagList = BEOPDataAccess.getInstance().getAllTagsFromEnv()
    if tagList is None:
        return jsonify(dict(err=1, msg="获取失败", data=[]))
    return jsonify(dict(err=0, msg="获取成功", data=tagList))

@bp_env.route("/getEnvByTags", methods=["POST"])
def get_env_by_tags():
    try:
        rcv = request.get_json()
        tags = rcv.get("tags") if rcv.get("tags") is not None else None
        if tags is None:
            return jsonify(dict(err=1, msg="tags不能为空", data=[]))
        if not isinstance(tags, list):
            return jsonify(dict(err=1, msg="tag必须为一个列表", data=[]))

        proList = []
        for tag in tags:
            try:
                if len(str(tag)):
                    proList.append(str(tag))
            except:
                pass

        proList = list(set(proList))

        if not len(proList):
            return jsonify(dict(err=1, msg="未发现tags", data=[]))

        envList = BEOPDataAccess.getInstance().getEnvByTags(proList)

        if envList is None:
            return jsonify(dict(err=1, msg="数据库繁忙，请重试", data=[]))

        res = list()
        if len(envList):
            for item in envList:
                res.append({
                    "id": item[0],
                    "type": item[1],
                    "enabled": item[2],
                    "name": item[3],
                    "description": item[4],
                    "tags": item[5].split(";"),
                    "creator": item[6],
                    "createtime": item[7].strftime("%Y-%m-%d %H:%M:%S")
                })

        return jsonify(dict(err=0, msg="获取成功", data=res))

    except Exception as e:
        logging.error("ERROR in /env/getEnvByTags: %s" % e.__str__())
        return jsonify(dict(err=1, msg="获取失败", data=[]))

