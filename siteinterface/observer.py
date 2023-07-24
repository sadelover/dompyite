# -*- coding: utf-8 -*-
"""
Routes and views for the flask application.
"""

from siteinterface import app
from siteinterface.BEOPDataAccess import *
from siteinterface.BEOPSqliteAccess import *
from siteinterface.views import *
from flask import Flask,request,session,g,make_response,redirect,url_for,abort,render_template,send_file,flash,json,jsonify
import mysql.connector
from datetime import datetime
import os
import json
import re 
import time
import hashlib
from ctypes import *
import xlrd
import xlwt
import io
import mimetypes
from flask import Response
from .sqlite_manager import SqliteManager
from siteinterface.mod_point.utils import DeepLogicPointTool
from siteinterface.utils import ProcessManager
from openpyxl import Workbook


def format_result(is_success, msg='', data=None):
    ''' 格式化输出结果 '''
    result = None
    if is_success:
        result = {"status": 'OK', "err": 0}
    else:
        result = {"status": 'ERROR', "err":1}
    if msg:
        result['msg'] = msg
    if data != None:
        result['data'] = data
    return jsonify(result)

@app.route('/login')
def home():
    """Renders the home page."""
    return render_template('debugTool.html')
    # return render_template('test.html')

@app.route('/debug')
def debug():
    # rv = redirect(url_for('homes'))
    #return  rv
    return render_template('debugTool.html')

@app.route('/search')
def search():
    return render_template('debugTool.html')

@app.route('/set_realtimedata_from_site', methods=['POST'])
def set_realtimedata_from_site():
    data = request.get_json()
    return jsonify(dict(err=0, msg='', data={}))

#mango added 2014-11-26
@app.route('/observer/account/addMember/<userId>/<roleId>/<projectId>')
def addMember(userId, roleId, projectId):
    return json.dumps(BEOPDataAccess.getInstance().add_member(userId, roleId, projectId), ensure_ascii=False)


#mango added 2014-11-26
@app.route('/observer/account/removeMember/<userId>/<projectId>')
def remove_member(userId,projectId):
    return json.dumps(BEOPDataAccess.getInstance().remove_member(userId, projectId), ensure_ascii=False)

#mango added 2014-11-27
@app.route('/observer/account/resetLevel/<memberId>/<roleId>/<projectId>')
def reset_level(memberId,roleId,projectId):
    return json.dumps(BEOPDataAccess.getInstance().reset_level(memberId,roleId,projectId), ensure_ascii=False)

@app.route('/observersq')
def homeSQ():
    """Renders the home page."""
    return render_template('indexObserver_SQ.html',
        title='Home Page',
        year=datetime.now().year,)

#mango added 2014-11-27
@app.route('/observer/account/resetPassword',methods=['POST'])
def reset_password():
    print('reset_password')
    rdata = request.get_json()
    return BEOPDataAccess.getInstance().reset_password(rdata)

#mango added 2014-11-27
@app.route('/observer/account/getDetails/<userId>')
def get_details(userId):
    return BEOPDataAccess.getInstance().get_details(userId)

#mango added 2014-11-27
@app.route('/observer/account/resetUserInfo', methods=['POST'])
def reset_user_info():
    print('reset_user_info')
    rdata = request.get_json()
    user_id = rdata.get('id')
    user_name = rdata.get('name')
    user_mail = rdata.get('mail')
    return BEOPDataAccess.getInstance().reset_user_info(user_id. user_name, user_mail)

#mango added 2014-11-27
@app.route('/observer/account/forgetPassword/<userId>')
def forget_pass(userId):
    return BEOPDataAccess.getInstance().forget_pass(userId)

#mango added 2014-11-28 @todo
@app.route('/observer/account/logout')
def logout():
    return None


@app.route('/get_plant/<proj>')
@app.route('/get_plant/<proj>/<pageid>')
def get_plant(proj=1, pageid=10000158):
    rv = BEOPSqliteAccess.getInstance('POST get_plant post interface').getPlantPageContent_With_Redis(pageid)
    return json.dumps(rv)

@app.route('/get_plant_DebugTool/<pageid>')
def get_plant_DebugTool(pageid=10000158):
    rv = BEOPSqliteAccess.getInstance('POST get_plant post interface').getPlantPageContentDebugTool_With_Redis(pageid)
    return json.dumps(rv)

@app.route('/getSystemEquipmentPage', methods=["POST"])
def get_system_equipment_page():
    rv = dict(buttons=[],
              texts=[],
              images=[],
              equipments=[],
              animationImages=[],
              animationList=[],
              timePickers=[],
              rulers=[],
              pipelines=[],
              charts=[],
              page=[],
              tempDistributions=[],
              gages=[],
              checkboxs=[],
              rects=[],
              customControls=[])
    try:
        req_json = request.get_json()

        pageId = req_json.get("templatePageId") if req_json.get("templatePageId") is not None else None
        dictParam = req_json.get("placeHolder") if req_json.get("placeHolder") is not None else None
        templateFileNameWithoutExt = req_json.get("templateFileName") if req_json.get("templateFileName") is not None else "template"

        if pageId == None or dictParam == None or templateFileNameWithoutExt == None:
            return json.dumps(rv)

        rv = BEOPSqliteAccess.getInstance('POST getSystemEquipmentPage').getSystemEquipmentPage(pageId, dictParam, templateFileNameWithoutExt)
        return json.dumps(rv)

    except Exception as e:
        strLog = "ERROR in get_system_equipment_page: %s" % e.__str__()
        logging.error(strLog)
        return json.dumps(rv)

@app.route("/getPageType/<proj>/<pageid>")
def get_page_type(proj=1, pageid=10000158):
    rv = BEOPSqliteAccess.getInstance('GET getPageType').getPageType(pageid)
    return json.dumps(rv)

@app.route("/debug/setProject", methods=["POST"])
def set_project():
    data = request.get_json()
    return jsonify(dict(err=0, msg="", data=""))

@app.route('/get_plant_dynamic')
@app.route('/get_plant_dynamic/<pageid>')
def get_plant_dynamic(pageid=10000158):
    rv = BEOPSqliteAccess.getInstance('GET get_plant_dynamic').getPlantDynamic(pageid)
    return json.dumps(rv)

@app.route('/logic/setDrawing', methods=["POST"])
def set_drawing():
    res = None
    try:
        req = request.get_json()
        of_logic_id = req.get("of_logic_id")
        content = req.get("content")
        if not isinstance(of_logic_id, str) or not isinstance(content, str):
            return jsonify(dict(err=0, msg="输入量数据类型有误", data=res))

        domdb_4db = ''
        if app.config.get("USE_4DB_FILE_FORMAT"):
            domdb_4db = app.config.get("USE_4DB_NAME")
            if not os.path.exists(domdb_4db):
                print(domdb_4db + ' file not existing!')
                return list()

        cur_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())

        with SqliteManager(domdb_4db) as dbm:
            str_sql1 = '''REPLACE INTO logic_drawing (of_logic_id, content, version_time)
                            VALUES(?, ?, ?);'''
            dbm.exec_none_query(str_sql1, (of_logic_id, content, cur_time))

        return jsonify(dict(err=0, msg="写入成功", data=None))

    except Exception as e:
        logging.error("setDrawing error: %s" % e)
        return jsonify(dict(err=1, msg="写入失败", data=None))


@app.route('/logic/getDrawing')
def get_drawing():
    res = dict()
    of_logic_id = request.args.get('of_logic_id', None)
    try:
        domdb_4db = ''
        if app.config.get("USE_4DB_FILE_FORMAT"):
            domdb_4db = app.config.get("USE_4DB_NAME")
            if not os.path.exists(domdb_4db):
                print(domdb_4db + ' file not existing!')
                return list()

        with SqliteManager(domdb_4db) as dbm:
            str_sql = '''SELECT                            
                            of_logic_id,
                            content,
                            version_time
                        FROM
                            logic_drawing
                        WHERE
                            of_logic_id=?
                        ORDER BY
                            version_time DESC
                            LIMIT 1'''
            arr = dbm.exec_query(str_sql, (of_logic_id,))
        if len(arr) == 0:
            return jsonify(dict(err=0, msg="记录为空", data=res))
        else:
            res = dict(of_logic_id=arr[0].get("of_logic_id").decode("utf-8"), content=arr[0].get("content").decode("utf-8"), version_time=arr[0].get("version_time").decode("utf-8"))
            return jsonify(dict(err=0, msg="", data=res))

    except Exception as e:
        logging.error("getDrawing error: %s" % e)
        return jsonify(dict(err=0, msg="获取失败", data=res))


@app.route('/get_plant_pagedetails/<proj>')
@app.route('/get_plant_pagedetails/<proj>/<userId>')
def get_plant_pagedetails(proj=1, userId=-1):
    s3dbname = BEOPDataAccess.getInstance().getProjS3db(proj)
    bMenuGrouped = BEOPSqliteAccess.getInstance('GET get_plant_pagedetails').GetMenuGroupEnabled(s3dbname)

    nUserRight = 99
    try:
        userId = int(userId)
        if userId is None:
            nUserRight = 0
        elif userId < 0:
            nUserRight = 99
        # 当userId为9999时，是访客，访客权限给1
        elif userId==9999:
            nUserRight = 1
        elif userId==0:
            nUserRight = 99
        elif userId >= 0:
            uInfo = BEOPDataAccess.getInstance().get_user_info(int(userId))
            if uInfo is None:
                nUserRight = 0
            else:
                nUserRight = uInfo.get('userofrole')
    except Exception as e:
        logging.error('ERROR in get_plant_pagedetails:%s' % (e.__str__()))
        pass

    try:
        nUserRight = int(nUserRight)
    except:
        nUserRight =99

    if not bMenuGrouped:
        rv = BEOPSqliteAccess.getInstance().getPlantPageDetails(s3dbname,None,  nUserRight)

        return json.dumps(rv, ensure_ascii=False)
    else:
        dataWithGroupInfo = dict(GroupMenuEnable= True, GroupList = [])

        groupList = BEOPSqliteAccess.getInstance().getPlantPageGroups(s3dbname)
        for group in groupList:
            rv = BEOPSqliteAccess.getInstance().getPlantPageDetails(s3dbname, group['id'], nUserRight)
            if rv:
                dataWithGroupInfo['GroupList'].append(dict(groupId= group['id'], groupName= group['name'], pageList = rv))
        return json.dumps(dataWithGroupInfo, ensure_ascii=False)


@app.route('/observer/account/sendInvitation',methods=['POST'])
def sendInvitation():
    reqdata = request.get_json()
    mail = reqdata.get('mail')
    prj_id = reqdata['prjID']
    level = reqdata['level']
    url = reqdata['serverURL']
    iname = reqdata['iname']
    imail = reqdata['imail']
    return BEOPDataAccess.getInstance().sendInvitation(mail,  prj_id, level,url,iname,imail)

def ProcessMail(inputMail):  
    isMatch = bool(re.match(r"^[a-zA-Z](([a-zA-Z0-9]*\.[a-zA-Z0-9]*)|[a-zA-Z0-9]*)[a-zA-Z]@([a-z0-9A-Z]+\.)+[a-zA-Z]{2,}$", inputMail,re.VERBOSE))  
    return isMatch


@app.route('/observer/account/login',methods=['POST'])
def login_():
    return json.dumps(BEOPDataAccess.getInstance().loginTask_(request.get_json()), ensure_ascii=False)


@app.route('/register',methods=['POST'])
def regist():

    data = request.get_json()
    name = data["name"]
    psw = data["password"]
   # token = data["token"]
    res = BEOPDataAccess.getInstance().regist(name,psw) #,token)
    return  res

@app.route('/register_new_user',methods=['POST'])
def register_new_user():
    data = request.get_json()
    username = data["username"]
    password = data["password"]

    res = BEOPDataAccess.getInstance().register_new_user(username, password)
    return res


@app.route('/invite_to_register/<token>')
def invite_to_register(token):
    userName = BEOPDataAccess.getInstance().getNameByToken(token)
    if BEOPDataAccess.getInstance().getUserStatus(userName) =='registered':
        return render_template('page_not_found.html'),404

    if len(userName) == 0:
        return render_template('page_not_found.html'), 404
    else :
        return render_template('register.html', name = userName,)

@app.route('/analysis/get_point_info_from_s3db', methods=['POST'])
def get_point_info_from_s3db():
    data = request.get_json()
    strPointNameList = data.get('pointList', [])
    if strPointNameList is None:
        return format_result(False, 'pointList param need in body', '')
    if not isinstance(strPointNameList, list):
        return format_result(False, 'pointList must be array', strPointNameList)
    # if not len(strPointNameList):
    #     return format_result(False, 'pointList is empty', '')

    ptNameList = []
    for pointName in strPointNameList:
        strName = pointName
        if pointName.find(" ") > -1:

            strName = None
            ptList = re.findall("<%.*?%>", pointName)
            if not len(ptList):
                ptList = re.findall('".*?"', pointName)

            if len(ptList):
                strName = ptList[0].replace("<", "")
                strName = strName.replace(">", "")
                strName = strName.replace("%", "")
                strName = strName.replace("'", "")
                strName = strName.replace('"', "")

        if strName:
            ptNameList.append(strName)

    result = BEOPSqliteAccess.getInstance().getPointInfoFromS3db(ptNameList)
    rvPointValueList = BEOPDataAccess.getInstance().getInputTable(ptNameList)[0]

    realtimeValueList = list()
    try:
        for obj in rvPointValueList:
            pointName = obj.get("name")
            obj["description"] = result.get(pointName).get("description")
            realtimeValueList.append(obj)
    except:
        pass

    result['realtimeValue'] = realtimeValueList

    # try:
    #     result['realtimeValue'] = BEOPDataAccess.getInstance().getInputTable(strPointNameList)[0]
    # except:
    #     pass

    return format_result(True, '获取通讯点列表成功', result)

@app.route('/analysis/get_all_points_info_by_type', methods=['POST'])
def get_points_info_by_type():
    data = request.get_json()
    strSourceType = data.get('sourceType')

    result = BEOPSqliteAccess.getInstance().getPointInfoFromS3db(None, strSourceType)
    return format_result(True, '获取通讯点列表成功', result)



'''
查找点名
'''
@app.route('/point/findByKeyword', methods=['POST'])
def point_find_by_key():
    data =  request.get_json()
    strKeyword = data.get('keyword')
    ret = BEOPSqliteAccess.getInstance().findPointListByKeyList_With_Redis(strKeyword, 1)

    return json.dumps( ret, ensure_ascii=False)


'''
获取点表列表
'''
@app.route('/point/getFaultPoints', methods=['POST'])
def get_fault_points():
    try:
        keyword = request.args.get('keyword', None)
        result = BEOPSqliteAccess.getInstance().getFaultPointListFromS3db_With_Redis()
    except:
        return format_result(False, '获取通讯点列表失败', [])

    return format_result(True, '获取通讯点列表成功', result)

'''
获取点表列表
'''
@app.route('/analysis/get_pointList_from_s3db')
@app.route('/analysis/get_pointList_from_s3db/<int:page_index>/<int:page_size>')
@app.route('/analysis/get_pointList_from_s3db/<int:projId>/<int:page_index>/<int:page_size>')
def get_pointList_from_s3db(page_index=1, page_size=50, projId=0):
    try:
        keyword = request.args.get('keyword', None)
        result = BEOPSqliteAccess.getInstance().getPointListFromS3db_With_Redis(page_index, page_size, keyword, projId)
    except:
        return format_result(False, '获取通讯点列表失败', [])

    return format_result(True, '获取通讯点列表成功', result)

'''
删除点表，POST提交的数据格式为：{pointList:[]}
'''
@app.route('/analysis/delete_point', methods=['POST'])
@app.route('/analysis/delete_point/<int:projId>', methods=['POST'])
def delete_point(projId=0):
    reqData = request.get_json()
    result = BEOPSqliteAccess.getInstance().deletePoint(reqData["pointList"], projId)
    return format_result(result['success'], '', result['recc'])


@app.route('/analysis/start', methods=['POST'])
def analysis_start():
    return json.dumps({dict(err=1,msg='not supported from new verion')})

@app.route('/analysis/generalRegressor', methods=['POST'])
def generalRegressor():
    # dll = cdll.LoadLibrary('beopWeb/lib/DataAnalysis.dll')
    # dll.GeneralRegressor.restype = c_char_p
    # szPara = create_string_buffer(json.dumps(request.get_json()).encode(encoding="utf-8"))
    # strResult = dll.GeneralRegressor(szPara)
    # return c_char_p(strResult).value.decode()
    return json.dumps({dict(err=1, msg='not supported from new verion')})

@app.route('/analysis/generalPredictor', methods=['POST'])
def generalPredictor():
    # dll = cdll.LoadLibrary('beopWeb/lib/DataAnalysis.dll')
    # dll.GeneralPredictor.restype = c_char_p
    # szPara = create_string_buffer(json.dumps(request.get_json()).encode(encoding="utf-8"))
    # strResult = dll.GeneralPredictor(szPara)
    # return c_char_p(strResult).value.decode()
    return json.dumps({dict(err=1, msg='not supported from new verion')})

@app.route("/getPngImageList")
def get_png_image_list():
    try:
        items = BEOPSqliteAccess.getInstance().getAutoReportTemplateIdListFileFrom4DB()

        pngList = []
        for item in items:
            fileName = item.get("file_name", None)
            if not isinstance(fileName, str):
                continue

            if fileName.endswith(".png") and fileName != "logo_small.png":
                pngList.append(fileName)

        return json.dumps((dict(err=0, msg="获取成功", data=pngList)))
    except Exception as e:
        strError = "获取失败：%s" % e.__str__()
        return json.dumps(dict(err=1, msg=strError, data=[]))

@app.route('/observer/report/check/<project>/<project_id>/<version>')
def checkReport(project, project_id, version):

    return json.dumps({dict(err=1, msg='not supported from new verion')})


@app.errorhandler(404)
def page_not_found(error):
    return render_template('404.html'), 404


@app.route('/pointTable/deepLogic/importPointTable', methods=['POST'])
def import_point_table_deeplogic():
    pointList = []
    strError = ""
    pointTable = request.files.get("file")
    if not pointTable:
        return jsonify(dict(err=1, msg='未发现点表文件', data=False))

    file_name = pointTable.filename
    if not (file_name.endswith('xlsx') or file_name.endswith('xls')):
        return jsonify(dict(err=1, msg='只支持.xlsx文件', data=False))

    static_folder = app.static_folder
    temp_folder = os.path.join(static_folder, "files", "temp")
    saveFilePath = os.path.join(temp_folder, file_name)
    if os.path.exists(saveFilePath):
        os.remove(saveFilePath)
    pointTable.save(saveFilePath)

    xlsxWorkBook = xlrd.open_workbook(saveFilePath)
    dpTool = DeepLogicPointTool()

    if not dpTool.is_deeplogic_point_table(xlsxWorkBook):
        return jsonify(dict(err=1, msg='检测到点表为非deeplogic点表，退出', data=False))

    dpTool.process_system_sheet(xlsxWorkBook)
    dpTool.process_equip_sheet(xlsxWorkBook)
    dpTool.process_point_sheet(xlsxWorkBook)
    dpTool.prepare_bacnet_point_list()
    dpTool.prepare_modbus_point_list()

    errMessageList = dpTool._lErr
    errMessageList = list(set(errMessageList))

    strError += "\n".join(errMessageList)

    pointList.extend(dpTool._lBacnetPoint)
    pointList.extend(dpTool._lModbusPoint)

    if not len(pointList):
        return jsonify(dict(err=1, msg="点表为空", data=False))

    """清除点表"""
    bCleanSuc = BEOPSqliteAccess.getInstance().clearPointList()
    if not bCleanSuc:
        strError += "\n清除点表失败"
        return jsonify(dict(err=1, msg=strError, data=False))

    """插入点表"""
    dInsertResult = BEOPSqliteAccess.getInstance().insertPointList(pointList, nProjSrc=1)

    """将最新的点表更新至redis"""
    BEOPSqliteAccess.getInstance().updateAllPointInfoIntoRedis()

    """重启domcore"""
    bSucKill = ProcessManager.getInstance().killProcess("domcore.exe")
    time.sleep(2)
    bSucStart = False
    if bSucKill:
        bSucStart = ProcessManager.getInstance().startProcess("domcore.exe", app.config['CORE_PATH'])

    bCoreRestart = bSucKill and bSucStart

    dInsertResult["success"] = dInsertResult.get("success") and bCoreRestart
    if not bCoreRestart:
        strError += "domcore自动重启失败，请手动重启domcore"

    if not dInsertResult.get("success"):
        strError += "\n" + dInsertResult.get("msg")

    if strError:
        log_info_to_file("deeplogic_point_%s.log" % datetime.now().strftime("%Y-%m-%d"), strError)
    try:
        os.remove(saveFilePath)
    except:
        pass
    return jsonify(dict(err=0, msg=strError, data=True))


@app.route('/pointTable/importPointTable', methods=['POST'])
@app.route('/pointTable/importPointTable/<int:projId>', methods=['POST'])
def import_point_table():
    errorList = []
    bSuc = False

    pointTable = request.files.get("file")
    if not pointTable:
        print("no pointTable captured")
        return json.dumps(dict(err=1, msg='no pointTable captured', data={}))

    file_name = pointTable.filename
    if not (file_name.endswith('xlsx') or file_name.endswith('xls')):
        print("only xlsx file supported")
        return json.dumps(dict(err=1, msg='only xlsx file supported', data={}))

    """
    nType:
    0-导入点表时直接覆盖；
    1-导入点表时若excel中的非vpoint与既有vpoint点名冲突，则用excel中的非vpoint替换既有的vpoint
    """
    strType = request.form.get("type", None)
    nType = 0
    if strType == None:
        nType = 0
    elif is_int_digit(strType):
        nType = int(strType)
    elif isinstance(strType, int):
        nType = strType

    if nType not in [0, 1]:
        nType = 0

    static_folder = app.static_folder
    temp_folder = os.path.join(static_folder, "files", "temp")
    saveFilePath = os.path.join(temp_folder, file_name)
    if os.path.exists(saveFilePath):
        os.remove(saveFilePath)
    pointTable.save(saveFilePath)

    xlsxWorkBook = xlrd.open_workbook(saveFilePath)

    sheet = xlsxWorkBook.sheet_by_index(0)
    if sheet.nrows <= 0:
        return jsonify(dict(err=1, msg="点表为空", data=False))

    """整理excel表"""
    allPointNameList = []
    allIdList = []
    dAllPointMap = {}
    dLegalPointIdMap = {}
    dIllegalPointIdMap = {}
    dDupPointMap = {}
    for idx in range(0, sheet.nrows):
        if idx == 0:
            continue

        rowValueList = sheet.row_values(idx)
        if not len(rowValueList):
            errorList.append(dict(name="", detail="点表中发现空行，但这不影响导入，已过滤"))
            continue

        if len(rowValueList) < 2:
            errorList.append(dict(name="", detail="点表中发现点位信息不完整的行，但这不影响导入，已过滤"))
            continue

        if not isinstance(rowValueList[1], str):
            errorList.append(dict(name="", detail="点表中发现点名不为字符串的行，但这不影响导入，已过滤"))
            continue

        if not len(rowValueList[1]):
            errorList.append(dict(name="", detail="点表中发现点名为空字符串的行，但这不影响导入，已过滤"))
            continue

        if re.match(r"^[0-9]{1}$", rowValueList[1][0]):
            errorList.append(dict(name=rowValueList[1], detail="点名({name})首字母为数字，不符合标准点名命名规范，已过滤".format(name=rowValueList[1])))
            continue

        if rowValueList[1].find(" ") >= 0:
            errorList.append(dict(name=rowValueList[1], detail="点名({name})中包含空格，不符合标准点名命名规范，已过滤".format(name=rowValueList[1])))
            continue

        if not re.match(r"^[a-zA-Z0-9_]*$", rowValueList[1]):
            errorList.append(dict(name=rowValueList[1], detail="点名({name})中包含特殊字符，不符合标准点名命名规范，已过滤".format(name=rowValueList[1])))
            continue

        id = rowValueList[0]
        pointNameRaw = rowValueList[1]

        if pointNameRaw in allPointNameList:
            if pointNameRaw not in dDupPointMap.keys():
                dDupPointMap.update({pointNameRaw: 0})

            dDupPointMap[pointNameRaw] += 1
            pointName = "{rawName}_Dup{n}".format(rawName=pointNameRaw, n=dDupPointMap[pointNameRaw])
            errorList.append(dict(name=pointNameRaw, detail="发现重复点名：{rawName}，已重命名为：{newName} (不影响导入)".format(rawName=pointNameRaw, newName=pointName)))
        else:
            pointName = pointNameRaw

        if not is_int_digit(id):  # 点位ID不合法
            dIllegalPointIdMap.update({pointName: None})
            dAllPointMap.update({pointName: [None, pointName] + rowValueList[2:]})
        else:  # 点位ID合法
            nId = int(id)
            if nId in allIdList:
                dIllegalPointIdMap.update({pointName: None})
            else:
                dLegalPointIdMap.update({pointName: nId})
                allIdList.append(nId)

            dAllPointMap.update({pointName: [int(id), pointName] + rowValueList[2:]})

        allPointNameList.append(pointName)

    """处理非法点位索引"""
    if dIllegalPointIdMap:
        nMaxId = 0
        legalIdList = list(dLegalPointIdMap.values())
        if len(legalIdList):
            nMaxId = max(legalIdList)

        nPointId = nMaxId + 1
        for key in dIllegalPointIdMap.keys():
            if not dAllPointMap.get(key, None):
                continue

            dAllPointMap[key][0] = nPointId
            nPointId += 1

    pointList = list(dAllPointMap.values())

    """nType为1：对点表进行补充操作"""
    if nType == 1:
        pointNameListToDeleteFromExisting = []
        toAddPointList = []
        dExistingPoint = BEOPSqliteAccess.getInstance().getPointInfoFromS3db([])
        nExistingIdList = [int(item.get("id")) for item in dExistingPoint.values()]
        nMaxIdOfExisting = max(nExistingIdList)

        for strPointName, lOnePointInfo in dAllPointMap.items():
            if dExistingPoint.get(strPointName):
                srcOfPointToAdd = lOnePointInfo[2]
                if srcOfPointToAdd != "vpoint" and dExistingPoint.get(strPointName).get("sourceType") == "vpoint":
                    pointNameListToDeleteFromExisting.append(strPointName)
                    toAddPointList.append(lOnePointInfo)
            else:
                toAddPointList.append(lOnePointInfo)

        nCur = nMaxIdOfExisting + 1
        for lOnePoint in toAddPointList:
            lOnePoint[0] = nCur
            nCur += 1

        bSuc, errSupplement = BEOPSqliteAccess.getInstance().supplementPoints(pointNameListToDeleteFromExisting, toAddPointList)
        if errSupplement:
            errorList.append(errSupplement)

    # nType为0: 对整个点表执行覆盖
    elif nType == 0:
        if not len(pointList):
            return jsonify(dict(err=1, msg="导入的excel表为空", data=False))

        """清除点表"""
        bCleanSuc = BEOPSqliteAccess.getInstance().clearPointList()
        if not bCleanSuc:
            return jsonify(dict(err=1, msg="点表清除失败", data=False))

        """插入点表"""
        dInsertResult = BEOPSqliteAccess.getInstance().insertPointList(pointList)
        bSuc = dInsertResult["success"]
        if dInsertResult["msg"]:
            errorList.append(dInsertResult["msg"])


    """重启domcore, 将最新的点表更新至redis """
    restart_domCore()
    BEOPSqliteAccess.getInstance().updateAllPointInfoIntoRedis()

    """导入点表过程中出现的问题反馈"""
    strErrMsg = "点表导入成功。" if bSuc else "点表导入失败。"
    errFileName = ""
    nErrFlag = 0  # 0 -没有问题；1  5条以内；2- 多条错误，txt文件供下载
    if len(errorList):
        strErrMsg += "另外在检查导入的excel表格时发现一些其他错误：\n"
        if len(errorList) <= 5:
            strErrMsg += ";".join([item.get("detail", "") for item in errorList])
            nErrFlag = 1
        else:
            for rootDir, dirs, fileNames in os.walk(temp_folder):
                for fileName in fileNames:
                    timeList = re.findall(r"[0-9]{4}-[0-9]{2}-[0-9]{2}-[0-9]{2}-[0-9]{2}-[0-9]{2}-[0-9]{6}", fileName)
                    if not len(timeList):
                        continue

                    strTime = timeList[0]
                    tTime = datetime.strptime(strTime, "%Y-%m-%d-%H-%M-%S-%f")
                    if (datetime.now() - tTime).total_seconds() > 6 * 3600:
                        try:
                            os.remove(os.path.join(rootDir, fileName))
                        except:
                            pass
            nErrFlag = 2
            errFileName = "pointImportErr_{time}.txt".format(time=datetime.now().strftime("%Y-%m-%d-%H-%M-%S-%f"))
            errMsgFilePath = os.path.join(temp_folder, errFileName)
            if os.path.exists(errMsgFilePath):
                try:
                    os.remove(errMsgFilePath)
                except:
                    pass
            with open(errMsgFilePath, "w", encoding="UTF8", errors="ignore") as fo:
                fo.write(";\n".join(["点名:{name}            描述:{detail}".format(name=item.get("name", ""), detail=item.get("detail", "")) for item in errorList]))
            strErrMsg += "错误超过5条，请下载查看具体错误描述"

    return jsonify(dict(err=0, msg=strErrMsg, data={"errFlag": nErrFlag, "errFileName": errFileName}))



'''
导出点表
'''
@app.route("/pointTable/exportPointTable")
@app.route("/pointTable/exportPointTable/<int:projId>")
def export_point_table(projId=0):
    strArrsystem = ["None", "HVAC", "Power", "Lighting", "CRAC"]
    strArrdevice = ["None", "Chiller", "Pump", "CT", "AHU", "VAV", "System"]
    strArrtype = ["None", "Power meter", "Thermal meter", "Tmperature", "Flow rate", "Pressure", "Amps", "Power",
                  "Frequency", "OnOff", "Alarm"]
    titles = ('pointindex','physicalid','source','remark','Unit','RWProperty','param1','param2','param3','param4','param5','param6','param7','param8','param9','param10','param11','param12','param13','param14','storecycle','customName','system','device','type')

    data = BEOPSqliteAccess.getInstance().getPointListForExport(projId)
    workbook = xlwt.Workbook('gbk')
    sheet = workbook.add_sheet('Sheet1')

    for tindex in range(len(titles)):
        sheet.write(0, tindex, titles[tindex])

    con = True
    for rindex in range(len(data)):
        row = data[rindex]
        for cindex in range(len(row)):
            col = row[cindex]
            if col is not None:
                if type(col) is bytes:
                    col = col.decode('gbk')
                if cindex == 5:
                    col = ('R', 'W')[col]
                elif cindex == 22:
                    col = strArrsystem[int(col) - 1]
                elif cindex == 23:
                    col = strArrdevice[int(col) - 1]
                elif cindex == 24:
                    col = strArrtype[int(col) - 1]
                elif cindex == 16 and int(col) == -9999:
                    col = ''
                sheet.write(rindex + 1, cindex, col)

    response = Response()
    response.status_code = 200

    output = io.BytesIO()
    workbook.save(output)
    response.data = output.getvalue()

    filename = 'pointlist.xls'
    mimetype_tuple = mimetypes.guess_type(filename)

    response.headers['Pragma'] = 'public'
    response.headers['Expires'] = '0'
    response.headers['Content-Type'] = mimetype_tuple[0]
    response.headers['Content-Transfer-Encoding'] = 'binary'
    response.headers['Content-Length'] = len(response.data)
    response.headers['Content-Disposition'] = 'attachment; filename="%s"' % filename

    return response


@app.route("/pointTable/exportPointTableV2")
def export_point_table_v2():
    projId = 0
    strArrsystem = ["None", "HVAC", "Power", "Lighting", "CRAC"]
    strArrdevice = ["None", "Chiller", "Pump", "CT", "AHU", "VAV", "System"]
    strArrtype = ["None", "Power meter", "Thermal meter", "Tmperature", "Flow rate", "Pressure", "Amps", "Power",
                  "Frequency", "OnOff", "Alarm"]
    titles = (
    'pointindex', 'physicalid', 'source', 'remark', 'Unit', 'RWProperty', 'param1', 'param2', 'param3', 'param4',
    'param5', 'param6', 'param7', 'param8', 'param9', 'param10', 'param11', 'param12', 'param13', 'param14',
    'storecycle', 'customName', 'system', 'device', 'type')

    try:
        pointList = BEOPSqliteAccess.getInstance().getPointListForExport(projId)

        book = Workbook()
        sheet = book.create_sheet("点表", 0)
        for idx, title in enumerate(titles):
            sheet.cell(row=1, column=idx+1, value=title)

        for rindex, row in enumerate(pointList):
            for cindex, col in enumerate(row):
                if col is not None:
                    if type(col) is bytes:
                        if cindex == 4:  # 单位要判断encoding类型
                            nCheck = check_bytes_encoding(col)
                            if nCheck == 0:
                                try:
                                    col = col.decode("utf8")
                                except:
                                    pass
                            elif nCheck == 1:
                                try:
                                    col = col.decode("gbk")
                                except:
                                    pass
                            else:
                                try:
                                    col = col.decode("gbk")
                                except:
                                    pass

                        else:
                            try:
                                col = col.decode('gbk')
                            except:
                                try:
                                    col = col.decode("UTF-8")
                                except:
                                    col = ""

                    if cindex == 5:
                        col = ('R', 'W')[col]
                    elif cindex == 22:
                        if is_digit(col):
                            col = strArrsystem[int(col) - 1]
                    elif cindex == 23:
                        if is_digit(col):
                            col = strArrdevice[int(col) - 1]
                    elif cindex == 24:
                        if is_digit(col):
                            col = strArrtype[int(col) - 1]
                    elif cindex == 16 and is_digit(col):
                        if int(col) == -9999:
                            col = ''

                    sheet.cell(row=rindex+2, column=cindex+1, value=col)

        fileName = "pointtable.xlsx"
        filesDir = os.path.join(app.static_folder, "files")
        if not os.path.exists(filesDir):
            os.mkdir(filesDir)

        filePath = os.path.join(filesDir, fileName)
        if os.path.exists(filePath):
            os.remove(filePath)

        book.save(filePath)

        return jsonify(dict(err=0, msg="导出成功", data=fileName))

    except Exception as e:
        logging.error("ERROR in export_point_table_v2: %s" % e.__str__())
        return jsonify(dict(err=1, msg="导出点表失败:%s" % e.__str__(), data=""))







